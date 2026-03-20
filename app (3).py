import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="VendorSync · Reconciliation",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');

:root {
    --bg: #0d0f14;
    --surface: #141720;
    --surface2: #1c2130;
    --border: #252c3d;
    --accent: #4f8eff;
    --accent2: #00d4aa;
    --warn: #ff8c42;
    --danger: #ff4d6d;
    --text: #e8ecf4;
    --muted: #7a8499;
    --matched: #00d4aa22;
    --matched-border: #00d4aa55;
    --unmatched: #ff4d6d22;
    --unmatched-border: #ff4d6d55;
    --partial: #ff8c4222;
    --partial-border: #ff8c4255;
}

html, body, [class*="css"] {
    font-family: 'DM Mono', monospace;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

.main { background: var(--bg) !important; }
.block-container { padding: 1.5rem 2rem !important; max-width: 1400px; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border);
}
[data-testid="stSidebar"] * { color: var(--text) !important; }

/* Header */
.recon-header {
    display: flex;
    align-items: center;
    gap: 1rem;
    margin-bottom: 2rem;
    padding-bottom: 1.5rem;
    border-bottom: 1px solid var(--border);
}
.recon-logo {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 800;
    background: linear-gradient(135deg, var(--accent), var(--accent2));
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: -0.03em;
}
.recon-subtitle {
    font-size: 0.75rem;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.12em;
}

/* Stat cards */
.stat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 1rem; margin-bottom: 1.5rem; }
.stat-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    position: relative;
    overflow: hidden;
}
.stat-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
}
.stat-card.matched::before { background: var(--accent2); }
.stat-card.unmatched::before { background: var(--danger); }
.stat-card.partial::before { background: var(--warn); }
.stat-card.total::before { background: var(--accent); }
.stat-label { font-size: 0.7rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.1em; }
.stat-value { font-family: 'Syne', sans-serif; font-size: 1.8rem; font-weight: 700; margin: 0.25rem 0; }
.stat-card.matched .stat-value { color: var(--accent2); }
.stat-card.unmatched .stat-value { color: var(--danger); }
.stat-card.partial .stat-value { color: var(--warn); }
.stat-card.total .stat-value { color: var(--accent); }
.stat-sub { font-size: 0.7rem; color: var(--muted); }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: var(--surface) !important;
    border-radius: 8px;
    padding: 4px;
    gap: 2px;
    border: 1px solid var(--border);
}
.stTabs [data-baseweb="tab"] {
    font-family: 'Syne', sans-serif !important;
    font-size: 0.8rem !important;
    font-weight: 600 !important;
    color: var(--muted) !important;
    background: transparent !important;
    border-radius: 6px !important;
    padding: 6px 16px !important;
}
.stTabs [aria-selected="true"] {
    background: var(--surface2) !important;
    color: var(--text) !important;
}

/* Dataframes */
[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }

/* Buttons */
.stButton > button {
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    background: linear-gradient(135deg, var(--accent), #3d6fcc) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.6rem 1.5rem !important;
    transition: all 0.2s !important;
}
.stButton > button:hover { opacity: 0.88 !important; transform: translateY(-1px); }

/* Download button */
.stDownloadButton > button {
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    background: var(--surface2) !important;
    color: var(--accent2) !important;
    border: 1px solid var(--accent2) !important;
    border-radius: 8px !important;
}

/* Upload area */
[data-testid="stFileUploader"] {
    background: var(--surface) !important;
    border: 1px dashed var(--border) !important;
    border-radius: 10px !important;
}

/* Section headers */
.section-tag {
    display: inline-block;
    font-family: 'Syne', sans-serif;
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.15em;
    padding: 3px 10px;
    border-radius: 4px;
    margin-bottom: 0.5rem;
}
.tag-matched { background: var(--matched); color: var(--accent2); border: 1px solid var(--matched-border); }
.tag-unmatched { background: var(--unmatched); color: var(--danger); border: 1px solid var(--unmatched-border); }
.tag-partial { background: var(--partial); color: var(--warn); border: 1px solid var(--partial-border); }
.tag-blue { background: #4f8eff22; color: var(--accent); border: 1px solid #4f8eff55; }

/* Info boxes */
.info-box {
    background: var(--surface);
    border: 1px solid var(--border);
    border-left: 3px solid var(--accent);
    border-radius: 6px;
    padding: 0.75rem 1rem;
    font-size: 0.8rem;
    color: var(--muted);
    margin-bottom: 1rem;
}

/* Alert */
[data-testid="stAlert"] { border-radius: 8px !important; }

/* Selectbox, multiselect */
[data-testid="stSelectbox"] > div, [data-testid="stMultiSelect"] > div {
    background: var(--surface) !important;
    border-color: var(--border) !important;
    border-radius: 8px !important;
}

/* Number input */
[data-testid="stNumberInput"] input {
    background: var(--surface) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
}

/* Expander */
[data-testid="stExpander"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# UTILITY FUNCTIONS
# ─────────────────────────────────────────────

def clean_doc_number(val):
    """Normalize document numbers for comparison."""
    if pd.isna(val):
        return ""
    s = str(val).strip().upper()
    s = re.sub(r'[\s\-_/]', '', s)
    return s

def extract_utr(val):
    """Extract UTR number from a string."""
    if pd.isna(val):
        return ""
    s = str(val).strip().upper()
    utr_pattern = re.search(r'[A-Z]{4}\d{18}|UTR[\s:]*([A-Z0-9]+)', s)
    if utr_pattern:
        return utr_pattern.group(0).replace(' ', '').replace(':', '')
    return s

def get_period(dt):
    """Return YYYY-MM string for period matching."""
    try:
        return pd.to_datetime(dt).strftime('%Y-%m')
    except:
        return ""

def round_amount(val, decimals=2):
    """Round to avoid floating-point noise."""
    try:
        return round(float(val), decimals)
    except:
        return 0.0

def is_debit_note(doc_type):
    """Detect if a row is a debit note."""
    if pd.isna(doc_type):
        return False
    s = str(doc_type).upper()
    return any(k in s for k in ['DEBIT NOTE', 'DN', 'DEBIT MEMO', 'DM', 'CREDIT MEMO', 'CM'])

def is_collection(doc_type, particulars=""):
    """Detect if a row is a payment/collection."""
    if pd.isna(doc_type):
        doc_type = ""
    s = str(doc_type).upper() + " " + str(particulars).upper()
    return any(k in s for k in ['PAYMENT', 'RECEIPT', 'COLLECTION', 'NEFT', 'RTGS', 'IMPS', 'CHEQUE', 'CHQ', 'TDS', 'BANK', 'UTR'])

# ─────────────────────────────────────────────
# LOAD & PARSE
# ─────────────────────────────────────────────

def load_vendor_ledger(file):
    """Parse vendor ledger Excel."""
    df = pd.read_excel(file, header=None)
    # Find header row
    header_row = None
    for i, row in df.iterrows():
        vals = [str(v).lower() for v in row if not pd.isna(v)]
        if any('doc' in v and 'date' in v for v in vals) or any('doc. date' in v for v in vals):
            header_row = i
            break
    if header_row is None:
        for i, row in df.iterrows():
            vals = [str(v).lower() for v in row if not pd.isna(v)]
            if any('debit' in v for v in vals) and any('credit' in v for v in vals):
                header_row = i
                break
    if header_row is None:
        header_row = 0

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    # Normalize column names
    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if 'doc' in cl and 'date' in cl:
            col_map[c] = 'doc_date'
        elif 'doc' in cl and ('no' in cl or 'num' in cl or 'type' not in cl) and 'date' not in cl and 'type' not in cl:
            col_map[c] = 'doc_no'
        elif 'doc' in cl and 'type' in cl:
            col_map[c] = 'doc_type'
        elif 'particular' in cl:
            col_map[c] = 'particulars'
        elif 'opening' in cl:
            col_map[c] = 'opening'
        elif 'debit' in cl:
            col_map[c] = 'debit'
        elif 'credit' in cl:
            col_map[c] = 'credit'
        elif 'closing' in cl or 'balance' in cl:
            col_map[c] = 'closing'
    df = df.rename(columns=col_map)

    needed = ['doc_date', 'doc_no', 'doc_type', 'debit', 'credit']
    for col in needed:
        if col not in df.columns:
            df[col] = np.nan

    df['doc_date'] = pd.to_datetime(df['doc_date'], errors='coerce', dayfirst=True)
    df['debit'] = pd.to_numeric(df['debit'], errors='coerce').fillna(0)
    df['credit'] = pd.to_numeric(df['credit'], errors='coerce').fillna(0)
    df['doc_no_clean'] = df['doc_no'].apply(clean_doc_number)
    df['period'] = df['doc_date'].apply(get_period)
    df = df[df['doc_no_clean'] != ''].reset_index(drop=True)
    df['_idx'] = df.index
    return df


def load_customer_ledger(file):
    """Parse customer ledger Excel."""
    df = pd.read_excel(file, header=None)
    header_row = None
    for i, row in df.iterrows():
        vals = [str(v).lower() for v in row if not pd.isna(v)]
        if any('document' in v for v in vals) and any('debit' in v or 'credit' in v for v in vals):
            header_row = i
            break
    if header_row is None:
        header_row = 0

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {}
    for c in df.columns:
        cl = c.lower()
        if 'date' in cl:
            col_map[c] = 'doc_date'
        elif 'type' in cl:
            col_map[c] = 'doc_type'
        elif ('no' in cl or 'detail' in cl or 'num' in cl) and 'date' not in cl and 'type' not in cl:
            col_map[c] = 'doc_no'
        elif 'debit' in cl:
            col_map[c] = 'debit'
        elif 'credit' in cl:
            col_map[c] = 'credit'
    df = df.rename(columns=col_map)

    needed = ['doc_date', 'doc_no', 'doc_type', 'debit', 'credit']
    for col in needed:
        if col not in df.columns:
            df[col] = np.nan

    df['doc_date'] = pd.to_datetime(df['doc_date'], errors='coerce', dayfirst=True)
    df['debit'] = pd.to_numeric(df['debit'], errors='coerce').fillna(0)
    df['credit'] = pd.to_numeric(df['credit'], errors='coerce').fillna(0)
    df['doc_no_clean'] = df['doc_no'].apply(clean_doc_number)
    df['period'] = df['doc_date'].apply(get_period)
    df = df[df['doc_no_clean'] != ''].reset_index(drop=True)
    df['_idx'] = df.index
    return df

# ─────────────────────────────────────────────
# RECONCILIATION ENGINE
# ─────────────────────────────────────────────

def run_reconciliation(vl, cl, tolerance=1.0):
    """
    Main reconciliation logic:
    1. Exclude fully-reversed invoices from vendor ledger
    2. Match invoices by document number
    3. Match debit notes by doc number, then by period+amount
    4. Match collections by UTR or period+amount
    5. Report all unmatched items
    """
    results = {
        'invoice_matched': [],
        'invoice_unmatched_vl': [],
        'invoice_unmatched_cl': [],
        'dn_matched': [],
        'dn_unmatched_vl': [],
        'dn_unmatched_cl': [],
        'collection_matched': [],
        'collection_unmatched_vl': [],
        'collection_unmatched_cl': [],
        'reversed_excluded': [],
    }

    vl = vl.copy()
    cl = cl.copy()
    vl['_matched'] = False
    cl['_matched'] = False

    # ── STEP 1: Identify & exclude fully-reversed invoices in vendor ledger ──
    # A reversal = same doc_no appears with both positive credit and a reversal entry
    vl_invoices_raw = vl[~vl['doc_type'].apply(is_debit_note) & ~vl['doc_type'].apply(lambda x: is_collection(x))].copy()
    doc_groups = vl_invoices_raw.groupby('doc_no_clean')
    reversed_docs = set()
    for doc_no, grp in doc_groups:
        total_credit = grp['credit'].sum()
        total_debit = grp['debit'].sum()
        # If debit == credit (net zero) → fully reversed
        if abs(total_credit - total_debit) < tolerance and len(grp) > 1:
            reversed_docs.add(doc_no)
            results['reversed_excluded'].extend(grp.to_dict('records'))
            vl.loc[grp.index, '_matched'] = True  # exclude from further matching

    # ── STEP 2: Match Invoices by Document Number ──
    vl_inv = vl[(~vl['_matched']) & (~vl['doc_type'].apply(is_debit_note)) & (~vl['doc_type'].apply(lambda x: is_collection(x, '')))].copy()
    cl_inv = cl[(~cl['_matched']) & (~cl['doc_type'].apply(is_debit_note)) & (~cl['doc_type'].apply(lambda x: is_collection(x, '')))].copy()

    for idx, vrow in vl_inv.iterrows():
        matches = cl_inv[(cl_inv['doc_no_clean'] == vrow['doc_no_clean']) & (~cl_inv['_matched'])]
        if not matches.empty:
            crow = matches.iloc[0]
            vl.at[idx, '_matched'] = True
            cl.at[crow['_idx'], '_matched'] = True
            cl_inv.at[crow.name, '_matched'] = True
            results['invoice_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Debit': vrow.get('debit', 0),
                'VL Credit': vrow.get('credit', 0),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Debit': crow.get('debit', 0),
                'CL Credit': crow.get('credit', 0),
                'Match Basis': 'Document Number',
                'Match Type': 'Invoice',
            })

    # ── STEP 3: Match Debit Notes ──
    vl_dn = vl[(~vl['_matched']) & (vl['doc_type'].apply(is_debit_note))].copy()
    cl_dn = cl[(~cl['_matched']) & (cl['doc_type'].apply(is_debit_note))].copy()

    for idx, vrow in vl_dn.iterrows():
        matched = False
        # 3a: by doc number
        doc_matches = cl_dn[(cl_dn['doc_no_clean'] == vrow['doc_no_clean']) & (~cl_dn['_matched'])]
        if not doc_matches.empty:
            crow = doc_matches.iloc[0]
            matched = True
            basis = 'Document Number'
        else:
            # 3b: by period + amount
            vamt = round_amount(vrow['debit'] + vrow['credit'])
            period_matches = cl_dn[
                (cl_dn['period'] == vrow['period']) &
                (abs(cl_dn['debit'] + cl_dn['credit'] - vamt) <= tolerance) &
                (~cl_dn['_matched'])
            ]
            if not period_matches.empty:
                crow = period_matches.iloc[0]
                matched = True
                basis = 'Period + Amount'

        if matched:
            vl.at[idx, '_matched'] = True
            cl.at[crow['_idx'], '_matched'] = True
            cl_dn.at[crow.name, '_matched'] = True
            results['dn_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Debit': vrow.get('debit', 0),
                'VL Credit': vrow.get('credit', 0),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Debit': crow.get('debit', 0),
                'CL Credit': crow.get('credit', 0),
                'Match Basis': basis,
                'Match Type': 'Debit Note',
            })

    # ── STEP 4: Match Collections by UTR or Period+Amount ──
    vl_col = vl[(~vl['_matched']) & (vl['doc_type'].apply(lambda x: is_collection(x, '')))].copy()
    cl_col = cl[(~cl['_matched']) & (cl['doc_type'].apply(lambda x: is_collection(x, '')))].copy()

    # Add UTR column
    vl_col['utr'] = vl_col.apply(lambda r: extract_utr(str(r.get('particulars', '')) + ' ' + str(r.get('doc_no', ''))), axis=1)
    cl_col['utr'] = cl_col.apply(lambda r: extract_utr(str(r.get('doc_no', '')) + ' ' + str(r.get('doc_type', ''))), axis=1)

    for idx, vrow in vl_col.iterrows():
        matched = False
        basis = ''
        # 4a: by UTR
        if vrow['utr']:
            utr_matches = cl_col[(cl_col['utr'] == vrow['utr']) & (cl_col['utr'] != '') & (~cl_col['_matched'])]
            if not utr_matches.empty:
                crow = utr_matches.iloc[0]
                matched = True
                basis = 'UTR Number'

        if not matched:
            # 4b: by period + amount
            vamt = round_amount(vrow['debit'] + vrow['credit'])
            amt_matches = cl_col[
                (cl_col['period'] == vrow['period']) &
                (abs(cl_col['debit'] + cl_col['credit'] - vamt) <= tolerance) &
                (~cl_col['_matched'])
            ]
            if not amt_matches.empty:
                crow = amt_matches.iloc[0]
                matched = True
                basis = 'Period + Amount'

        if matched:
            vl.at[idx, '_matched'] = True
            cl.at[crow['_idx'], '_matched'] = True
            cl_col.at[crow.name, '_matched'] = True
            results['collection_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Amount': vrow.get('debit', 0) + vrow.get('credit', 0),
                'VL UTR': vrow.get('utr', ''),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Amount': crow.get('debit', 0) + crow.get('credit', 0),
                'CL UTR': crow.get('utr', ''),
                'Match Basis': basis,
                'Match Type': 'Collection',
            })

    # ── STEP 5: Collect all unmatched ──
    unmatched_vl = vl[~vl['_matched']]
    unmatched_cl = cl[~cl['_matched']]

    for _, r in unmatched_vl.iterrows():
        entry = {
            'Doc No': r.get('doc_no', ''),
            'Date': r.get('doc_date', ''),
            'Type': r.get('doc_type', ''),
            'Particulars': r.get('particulars', ''),
            'Debit': r.get('debit', 0),
            'Credit': r.get('credit', 0),
            'Source': 'Vendor Ledger',
        }
        if is_debit_note(r.get('doc_type', '')):
            results['dn_unmatched_vl'].append(entry)
        elif is_collection(r.get('doc_type', '')):
            results['collection_unmatched_vl'].append(entry)
        else:
            results['invoice_unmatched_vl'].append(entry)

    for _, r in unmatched_cl.iterrows():
        entry = {
            'Doc No': r.get('doc_no', ''),
            'Date': r.get('doc_date', ''),
            'Type': r.get('doc_type', ''),
            'Debit': r.get('debit', 0),
            'Credit': r.get('credit', 0),
            'Source': 'Customer Ledger',
        }
        if is_debit_note(r.get('doc_type', '')):
            results['dn_unmatched_cl'].append(entry)
        elif is_collection(r.get('doc_type', '')):
            results['collection_unmatched_cl'].append(entry)
        else:
            results['invoice_unmatched_cl'].append(entry)

    return results

# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

def build_excel(results):
    output = BytesIO()
    wb = openpyxl.Workbook()

    COLORS = {
        'header_bg': '1C2130',
        'header_fg': 'E8ECF4',
        'matched': 'D4F5EE',
        'unmatched': 'FFE5EA',
        'partial': 'FFF3E5',
        'reversed': 'F0F0F0',
        'alt_row': 'F8F9FB',
        'border': 'D0D5E0',
    }

    thin = Side(style='thin', color=COLORS['border'])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, row=1, color='1C2130'):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            cell.fill = PatternFill(fill_type='solid', fgColor=color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 30

    def style_data(ws, start_row, end_row, ncols, row_color=None):
        for r in range(start_row, end_row + 1):
            fill_color = row_color if row_color else (COLORS['alt_row'] if r % 2 == 0 else 'FFFFFF')
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                cell.font = Font(name='Calibri', size=9)

    def auto_width(ws, min_w=10, max_w=40):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, max_len + 2))

    # ── Summary Sheet ──
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    summary_data = [
        ['Category', 'Matched', 'Unmatched (VL)', 'Unmatched (CL)', 'Total VL', 'Total CL'],
        ['Invoices',
         len(results['invoice_matched']),
         len(results['invoice_unmatched_vl']),
         len(results['invoice_unmatched_cl']),
         len(results['invoice_matched']) + len(results['invoice_unmatched_vl']),
         len(results['invoice_matched']) + len(results['invoice_unmatched_cl'])],
        ['Debit Notes',
         len(results['dn_matched']),
         len(results['dn_unmatched_vl']),
         len(results['dn_unmatched_cl']),
         len(results['dn_matched']) + len(results['dn_unmatched_vl']),
         len(results['dn_matched']) + len(results['dn_unmatched_cl'])],
        ['Collections',
         len(results['collection_matched']),
         len(results['collection_unmatched_vl']),
         len(results['collection_unmatched_cl']),
         len(results['collection_matched']) + len(results['collection_unmatched_vl']),
         len(results['collection_matched']) + len(results['collection_unmatched_cl'])],
        ['Reversed (Excluded)', len(results['reversed_excluded']), '', '', len(results['reversed_excluded']), ''],
    ]
    for r, row in enumerate(summary_data, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    style_header(ws, summary_data[0], row=1)
    style_data(ws, 2, len(summary_data), len(summary_data[0]))
    auto_width(ws)

    # ── Helper to write a sheet ──
    def write_sheet(title, data, color):
        if not data:
            return
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A2'
        df = pd.DataFrame(data)
        headers = list(df.columns)
        style_header(ws, headers, row=1, color=color)
        for r, (_, row) in enumerate(df.iterrows(), 2):
            for c, val in enumerate(row.values, 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(val, (pd.Timestamp, datetime)):
                    cell.value = val.strftime('%d-%b-%Y') if not pd.isna(val) else ''
                elif isinstance(val, float):
                    cell.value = round(val, 2)
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = val
        style_data(ws, 2, len(data) + 1, len(headers),
                   row_color=None)
        auto_width(ws)

    write_sheet('Inv - Matched', results['invoice_matched'], '1A6B45')
    write_sheet('Inv - Unmatched VL', results['invoice_unmatched_vl'], 'A32035')
    write_sheet('Inv - Unmatched CL', results['invoice_unmatched_cl'], 'A32035')
    write_sheet('DN - Matched', results['dn_matched'], '1A6B45')
    write_sheet('DN - Unmatched VL', results['dn_unmatched_vl'], 'A32035')
    write_sheet('DN - Unmatched CL', results['dn_unmatched_cl'], 'A32035')
    write_sheet('Collections - Matched', results['collection_matched'], '1A6B45')
    write_sheet('Collections - Unmatch VL', results['collection_unmatched_vl'], 'A32035')
    write_sheet('Collections - Unmatch CL', results['collection_unmatched_cl'], 'A32035')
    write_sheet('Reversed Excluded', results['reversed_excluded'], '555555')

    wb.save(output)
    return output.getvalue()

# ─────────────────────────────────────────────
# UI HELPERS
# ─────────────────────────────────────────────

def fmt_inr(val):
    try:
        return f"₹{val:,.2f}"
    except:
        return str(val)

def display_df(df, color='default'):
    if df.empty:
        st.info("No records in this category.")
        return
    # Format date columns
    for col in df.columns:
        if 'date' in col.lower() or 'Date' in col:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d-%b-%Y').fillna('')
    st.dataframe(df, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────

def main():
    # Header
    st.markdown("""
    <div class="recon-header">
        <div>
            <div class="recon-logo">⚖️ VendorSync</div>
            <div class="recon-subtitle">Vendor · Customer Ledger Reconciliation · For Indian CAs & CFOs</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")
        st.markdown("---")

        tolerance = st.number_input(
            "Amount Tolerance (₹)",
            min_value=0.0, max_value=100.0, value=1.0, step=0.5,
            help="Max difference allowed when matching by amount (e.g. rounding differences)"
        )

        st.markdown("---")
        st.markdown("### 📋 Matching Rules")
        st.markdown("""
        <div class="info-box">
        1. Invoices → Doc Number<br>
        2. Reversed invoices → Excluded<br>
        3. Debit Notes → Doc No → Period+Amount<br>
        4. Collections → UTR → Period+Amount<br>
        5. Remaining → Unmatched report
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### 🔍 Column Detection")
        st.markdown("""
        <div class="info-box" style="font-size:0.72rem;">
        Ledger columns are auto-detected. Ensure your Excel has standard column headers as per the format.
        </div>
        """, unsafe_allow_html=True)

    # File upload
    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<span class="section-tag tag-blue">VENDOR LEDGER</span>', unsafe_allow_html=True)
        vl_file = st.file_uploader("Upload Vendor Ledger (.xlsx / .xls)", type=['xlsx', 'xls'], key='vl')

    with col2:
        st.markdown('<span class="section-tag tag-blue">CUSTOMER LEDGER</span>', unsafe_allow_html=True)
        cl_file = st.file_uploader("Upload Customer Ledger (.xlsx / .xls)", type=['xlsx', 'xls'], key='cl')

    if not vl_file or not cl_file:
        st.markdown("""
        <div class="info-box" style="margin-top:2rem; border-left-color: #4f8eff;">
        📂 Upload both ledger files above to begin reconciliation. The engine will automatically detect columns,
        exclude reversed entries, and match invoices, debit notes, and collections using the configured rules.
        </div>
        """, unsafe_allow_html=True)
        return

    # Parse files
    with st.spinner("Parsing ledgers..."):
        try:
            vl = load_vendor_ledger(vl_file)
            cl = load_customer_ledger(cl_file)
        except Exception as e:
            st.error(f"Error reading files: {e}")
            return

    st.success(f"✅ Vendor Ledger: **{len(vl)}** rows  ·  Customer Ledger: **{len(cl)}** rows")

    # Preview toggle
    with st.expander("👁 Preview Parsed Data"):
        pc1, pc2 = st.columns(2)
        with pc1:
            st.markdown("**Vendor Ledger (first 10 rows)**")
            display_cols = [c for c in ['doc_date', 'doc_no', 'doc_type', 'particulars', 'debit', 'credit'] if c in vl.columns]
            st.dataframe(vl[display_cols].head(10), use_container_width=True, hide_index=True)
        with pc2:
            st.markdown("**Customer Ledger (first 10 rows)**")
            display_cols = [c for c in ['doc_date', 'doc_no', 'doc_type', 'debit', 'credit'] if c in cl.columns]
            st.dataframe(cl[display_cols].head(10), use_container_width=True, hide_index=True)

    # Run reconciliation
    if st.button("▶ Run Reconciliation", use_container_width=False):
        with st.spinner("Running reconciliation engine..."):
            results = run_reconciliation(vl, cl, tolerance=tolerance)
        st.session_state['results'] = results

    if 'results' not in st.session_state:
        return

    results = st.session_state['results']

    # Stats
    total_matched = len(results['invoice_matched']) + len(results['dn_matched']) + len(results['collection_matched'])
    total_unmatched = (len(results['invoice_unmatched_vl']) + len(results['invoice_unmatched_cl']) +
                       len(results['dn_unmatched_vl']) + len(results['dn_unmatched_cl']) +
                       len(results['collection_unmatched_vl']) + len(results['collection_unmatched_cl']))

    st.markdown(f"""
    <div class="stat-grid">
        <div class="stat-card matched">
            <div class="stat-label">Total Matched</div>
            <div class="stat-value">{total_matched}</div>
            <div class="stat-sub">Across all categories</div>
        </div>
        <div class="stat-card unmatched">
            <div class="stat-label">Unmatched Items</div>
            <div class="stat-value">{total_unmatched}</div>
            <div class="stat-sub">Needs review</div>
        </div>
        <div class="stat-card partial">
            <div class="stat-label">Reversed & Excluded</div>
            <div class="stat-value">{len(results['reversed_excluded'])}</div>
            <div class="stat-sub">Fully reversed invoices</div>
        </div>
        <div class="stat-card total">
            <div class="stat-label">Invoice Matches</div>
            <div class="stat-value">{len(results['invoice_matched'])}</div>
            <div class="stat-sub">DN: {len(results['dn_matched'])} · Coll: {len(results['collection_matched'])}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Download button
    excel_data = build_excel(results)
    st.download_button(
        label="⬇ Download Full Reconciliation Report (.xlsx)",
        data=excel_data,
        file_name=f"reconciliation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )

    st.markdown("---")

    # Detailed tabs
    tabs = st.tabs([
        "🧾 Invoices",
        "📝 Debit Notes",
        "💰 Collections",
        "🔁 Reversed",
        "⚠️ All Unmatched",
    ])

    with tabs[0]:
        st.markdown('<span class="section-tag tag-matched">MATCHED INVOICES</span>', unsafe_allow_html=True)
        display_df(pd.DataFrame(results['invoice_matched']))
        c1, c2 = st.columns(2)
        with c1:
            st.markdown('<span class="section-tag tag-unmatched">UNMATCHED — VENDOR LEDGER</span>', unsafe_allow_html=True)
            display_df(pd.DataFrame(results['invoice_unmatched_vl']))
        with c2:
            st.markdown('<span class="section-tag tag-unmatched">UNMATCHED — CUSTOMER LEDGER</span>', unsafe_allow_html=True)
            display_df(pd.DataFrame(results['invoice_unmatched_cl']))

    with tabs[1]:
        st.markdown('<span class="section-tag tag-matched">MATCHED DEBIT NOTES</span>', unsafe_allow_html=True)
        display_df(pd.DataFrame(results['dn_matched']))
        c1, c2 = st.columns(2)
        with c1:
            st.markdown('<span class="section-tag tag-unmatched">UNMATCHED — VENDOR LEDGER</span>', unsafe_allow_html=True)
            display_df(pd.DataFrame(results['dn_unmatched_vl']))
        with c2:
            st.markdown('<span class="section-tag tag-unmatched">UNMATCHED — CUSTOMER LEDGER</span>', unsafe_allow_html=True)
            display_df(pd.DataFrame(results['dn_unmatched_cl']))

    with tabs[2]:
        st.markdown('<span class="section-tag tag-matched">MATCHED COLLECTIONS</span>', unsafe_allow_html=True)
        display_df(pd.DataFrame(results['collection_matched']))
        c1, c2 = st.columns(2)
        with c1:
            st.markdown('<span class="section-tag tag-unmatched">UNMATCHED — VENDOR LEDGER</span>', unsafe_allow_html=True)
            display_df(pd.DataFrame(results['collection_unmatched_vl']))
        with c2:
            st.markdown('<span class="section-tag tag-unmatched">UNMATCHED — CUSTOMER LEDGER</span>', unsafe_allow_html=True)
            display_df(pd.DataFrame(results['collection_unmatched_cl']))

    with tabs[3]:
        st.markdown('<span class="section-tag tag-partial">REVERSED & EXCLUDED INVOICES</span>', unsafe_allow_html=True)
        st.caption("These invoices were fully reversed in the Vendor Ledger (Debit = Credit) and excluded from matching.")
        display_df(pd.DataFrame(results['reversed_excluded']))

    with tabs[4]:
        all_unmatched = []
        for item in results['invoice_unmatched_vl'] + results['dn_unmatched_vl'] + results['collection_unmatched_vl']:
            item['Ledger'] = 'Vendor'; all_unmatched.append(item)
        for item in results['invoice_unmatched_cl'] + results['dn_unmatched_cl'] + results['collection_unmatched_cl']:
            item['Ledger'] = 'Customer'; all_unmatched.append(item)
        st.markdown('<span class="section-tag tag-unmatched">ALL UNMATCHED ITEMS</span>', unsafe_allow_html=True)
        display_df(pd.DataFrame(all_unmatched))


if __name__ == "__main__":
    main()
