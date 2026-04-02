import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Ledger Reconciliation",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None,
    }
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');

/* ── Hide Streamlit chrome ── */
#MainMenu {visibility: hidden !important;}
footer {visibility: hidden !important;}
header {visibility: hidden !important;}
[data-testid="stToolbar"] {display: none !important;}
[data-testid="stDecoration"] {display: none !important;}
.viewerBadge_container__1QSob {display: none !important;}
.stDeployButton {display: none !important;}

:root {
    --bg:           #0d0f14;
    --surface:      #141720;
    --surface2:     #1a1d28;
    --border:       #252c3d;
    --accent:       #4f8eff;
    --accent2:      #00d4aa;
    --warn:         #ff8c42;
    --danger:       #ff4d6d;
    --text:         #e8ecf4;
    --muted:        #7a8599;
    --matched:      #0d2218;
    --matched-border:#1a6b45;
    --unmatched:    #2a0d14;
    --unmatched-border:#ff4d6d;
    --partial:      #2a1800;
    --partial-border:#ff8c42;
    --vl-color:     #4f8eff;
    --vl-bg:        #0d1628;
    --vl-border:    #1a3a6b;
    --cl-color:     #00d4aa;
    --cl-bg:        #0d2218;
    --cl-border:    #1a6b45;
}

html, body, [class*="css"] {
    font-family: 'Syne', sans-serif;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

.main { background: var(--bg) !important; }
.block-container { padding: 1.5rem 2rem !important; max-width: 1400px; }

[data-testid="stSidebar"] {
    background: #0d0f14 !important;
    border-right: 1px solid #252c3d;
}
[data-testid="stSidebar"] * { color: #e8ecf4 !important; }
[data-testid="stSidebar"] .stMarkdown p { color: #7a8599 !important; }
[data-testid="stSidebar"] input { background: #1a1d28 !important; color: #e8ecf4 !important; border-color: #252c3d !important; }

.recon-header {
    display: flex;
    align-items: center;
    gap: 1rem;
    margin-bottom: 1.5rem;
    padding: 1.25rem 1.5rem;
    border-radius: 12px;
    background: linear-gradient(135deg, #0d1628, #0d2218);
    border: 1px solid #252c3d;
}
.recon-logo {
    font-family: 'Syne', sans-serif;
    font-size: 1.8rem;
    font-weight: 800;
    color: #e8ecf4;
    letter-spacing: -0.02em;
}
.recon-subtitle {
    font-size: 0.72rem;
    color: #7a8599;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    margin-top: 2px;
}

.stat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 1rem; margin-bottom: 1.5rem; }
.stat-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1.1rem 1.25rem;
    position: relative;
    overflow: hidden;
    box-shadow: 0 2px 8px rgba(0,0,0,0.4);
}
.stat-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
}
.stat-card.matched::before { background: #1A6B45; }
.stat-card.unmatched::before { background: #A32035; }
.stat-card.partial::before { background: #B85C00; }
.stat-card.total::before { background: #1A3A6B; }
.stat-label { font-size: 0.68rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.1em; font-weight: 600; }
.stat-value { font-family: 'Syne', sans-serif; font-size: 1.9rem; font-weight: 700; margin: 0.2rem 0; color: var(--text); }
.stat-card.matched .stat-value  { color: #00d4aa; }
.stat-card.unmatched .stat-value{ color: #ff4d6d; }
.stat-card.partial .stat-value  { color: #ff8c42; }
.stat-card.total .stat-value    { color: #4f8eff; }
.stat-sub { font-size: 0.68rem; color: var(--muted); }

.stTabs [data-baseweb="tab-list"] {
    background: var(--surface) !important;
    border-radius: 8px;
    padding: 4px;
    gap: 2px;
    border: 1px solid var(--border);
}
.stTabs [data-baseweb="tab"] {
    font-family: 'Syne', sans-serif !important;
    font-size: 0.8rem !important;
    font-weight: 600 !important;
    color: var(--muted) !important;
    background: transparent !important;
    border-radius: 6px !important;
    padding: 6px 16px !important;
}
.stTabs [aria-selected="true"] {
    background: #1a1d28 !important;
    color: #4f8eff !important;
    border: 1px solid #252c3d !important;
}

[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; box-shadow: 0 1px 6px rgba(0,0,0,0.06); }

.stButton > button {
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    background: linear-gradient(135deg, #1a1d28, #252c3d) !important;
    color: #4f8eff !important;
    border: 1px solid #4f8eff !important;
    border-radius: 8px !important;
    padding: 0.55rem 1.4rem !important;
    transition: all 0.2s !important;
    box-shadow: 0 2px 8px rgba(79,142,255,0.15) !important;
}
.stButton > button:hover { opacity: 0.9 !important; transform: translateY(-1px) !important; }

.stDownloadButton > button {
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    background: var(--surface) !important;
    color: #00d4aa !important;
    border: 1px solid #00d4aa !important;
    border-radius: 8px !important;
}

[data-testid="stFileUploader"] {
    background: var(--surface) !important;
    border: 2px dashed var(--border) !important;
    border-radius: 10px !important;
}

.section-tag {
    display: inline-block;
    font-family: 'Inter', sans-serif;
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    padding: 4px 10px;
    border-radius: 4px;
    margin-bottom: 0.5rem;
}
.tag-matched   { background: var(--matched);   color: #1A6B45; border: 1px solid var(--matched-border); }
.tag-unmatched { background: var(--unmatched); color: #A32035; border: 1px solid var(--unmatched-border); }
.tag-partial   { background: var(--partial);   color: #B85C00; border: 1px solid var(--partial-border); }
.tag-blue      { background: var(--vl-bg);     color: #1A3A6B; border: 1px solid var(--vl-border); }
.tag-vl        { background: var(--vl-bg);     color: var(--vl-color); border: 1px solid var(--vl-border); }
.tag-cl        { background: var(--cl-bg);     color: var(--cl-color); border: 1px solid var(--cl-border); }

.info-box {
    background: var(--surface);
    border: 1px solid var(--border);
    border-left: 4px solid #1A3A6B;
    border-radius: 6px;
    padding: 0.85rem 1rem;
    font-size: 0.82rem;
    color: var(--text);
    margin-bottom: 1rem;
    line-height: 1.6;
}

[data-testid="stAlert"] { border-radius: 8px !important; }

[data-testid="stNumberInput"] input {
    background: var(--surface) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
}

[data-testid="stExpander"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 10px !important;
}

[data-testid="stTextInput"] input {
    background: var(--surface) !important;
    border-color: var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
    font-family: 'Syne', sans-serif !important;
}

/* Column mapping table */
.col-map-table { width:100%; border-collapse:collapse; font-size:0.8rem; margin-top:0.5rem; }
.col-map-table th { background:#1a1d28; color:#4f8eff; padding:7px 12px; text-align:left; border-bottom: 1px solid #252c3d; }
.col-map-table td { padding:6px 12px; border-bottom:1px solid var(--border); color:var(--text); }
.col-map-table tr:nth-child(even) td { background:var(--surface2); }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# UTILITY FUNCTIONS
# ─────────────────────────────────────────────

def clean_doc_number(val):
    if pd.isna(val):
        return ""
    s = str(val).strip().upper()
    s = re.sub(r'[\s\-_/]', '', s)
    return s

def extract_utr(val):
    if pd.isna(val):
        return ""
    s = str(val).strip().upper()
    utr_pattern = re.search(r'[A-Z]{4}\d{18}|UTR[\s:]*([A-Z0-9]+)', s)
    if utr_pattern:
        return utr_pattern.group(0).replace(' ', '').replace(':', '')
    return s

def get_period(dt):
    try:
        return pd.to_datetime(dt).strftime('%Y-%m')
    except:
        return ""

def round_amount(val, decimals=2):
    try:
        return round(float(val), decimals)
    except:
        return 0.0

def is_debit_note(doc_type):
    if pd.isna(doc_type):
        return False
    s = str(doc_type).upper()
    return any(k in s for k in ['DEBIT NOTE', 'DN', 'DEBIT MEMO'])

def is_reversal_type(doc_type):
    """
    Detect ONLY Complete Reversal entries.
    Saleable Return / Non-Saleable Return / Sales Return are treated as Credit Notes (not reversals).
    """
    if pd.isna(doc_type):
        return False
    s = str(doc_type).upper()
    return 'COMPLETE REVERSAL' in s

def is_credit_note(doc_type):
    """
    Credit Notes in VL — includes:
    - Credit Note / Credit Memo / Credit
    - Saleable Return / Non-Saleable Return (these are credit-type entries, NOT reversals)
    These are matched against Discount Debit Notes and PRN entries in CL.
    """
    if pd.isna(doc_type):
        return False
    s = str(doc_type).upper()
    return any(k in s for k in [
        'CREDIT NOTE', 'CREDIT MEMO', 'CREDIT',
        'SALEABLE RETURN', 'NON SALEABLE', 'NON-SALEABLE',
        'NONSALEABLE', 'SALE RETURN', 'SALES RETURN',
    ])

def is_discount_or_prn(doc_type, doc_no=""):
    """
    Detect CL entries that are Discount Debit Notes or PRN entries.
    These should be matched against VL Credit Notes.
    """
    if pd.isna(doc_type):
        doc_type = ""
    s = str(doc_type).upper()
    d = str(doc_no).upper()
    return any(k in s or k in d for k in [
        'DISCOUNT', 'DISC', 'DEBIT NOTE', 'DN',
        'PRN', 'PRICE REVISION', 'PRICE ADJ', 'REBATE',
        'ALLOWANCE', 'SCHEME', 'CLAIM'
    ])

def is_collection(doc_type, particulars=""):
    if pd.isna(doc_type):
        doc_type = ""
    s = str(doc_type).upper() + " " + str(particulars).upper()
    return any(k in s for k in ['PAYMENT', 'RECEIPT', 'COLLECTION', 'NEFT', 'RTGS', 'IMPS', 'CHEQUE', 'CHQ', 'TDS', 'BANK', 'UTR'])

def extract_ref_from_particulars(particulars):
    """Extract referenced invoice number from particulars column."""
    if pd.isna(particulars):
        return ""
    s = str(particulars).strip().upper()
    # Remove common prefix words
    s = re.sub(r'(REVERSAL OF|AGAINST|REF|REFERENCE|RETURN OF|CANCELLATION OF|REVERSED)\s*', '', s)
    # Clean up
    s = re.sub(r'[\s\-_/]', '', s)
    return s.strip()

# ─────────────────────────────────────────────
# LOAD & PARSE
# ─────────────────────────────────────────────

def _detect_header_row(df):
    """
    Scan rows to find the header row — the row where we see
    column-like labels (date, debit, credit, no, doc, etc).
    Works for ANY ledger format from any ERP system.
    Returns integer row index (0-based).
    """
    HEADER_KEYWORDS = ['date', 'debit', 'credit', 'doc', 'document',
                       'voucher', 'vch', 'amount', 'balance', 'no.',
                       'number', 'type', 'particular', 'narration']
    best_row = 0
    best_score = 0
    # Only scan first 20 rows
    for i in range(min(20, len(df))):
        try:
            row_vals = df.iloc[i].tolist()
        except Exception:
            continue
        score = 0
        for v in row_vals:
            try:
                s = str(v).lower().strip()
                if any(k in s for k in HEADER_KEYWORDS):
                    score += 1
            except Exception:
                continue
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _get_closing_from_raw(df):
    """
    Find closing balance = last non-null numeric value from
    a column whose name contains 'closing' or 'balance'.
    Works for any ledger format.
    """
    try:
        for c in df.columns:
            col_name = str(c).lower().strip()
            if 'closing' in col_name or ('balance' in col_name and 'opening' not in col_name):
                series = pd.to_numeric(df[c], errors='coerce').dropna()
                if not series.empty:
                    return float(series.iloc[-1])
    except Exception:
        pass
    return None


def _parse_ledger_raw(raw_df):
    """
    Given a raw DataFrame (no header), detect header row,
    set column names, capture closing balance from ALL rows
    (before doc_no filtering), and return (df_with_headers, closing_val).
    """
    header_row_idx = _detect_header_row(raw_df)

    # Extract column names from header row — safely convert each cell to string
    try:
        header_vals = raw_df.iloc[header_row_idx].tolist()
        col_names = [str(v).strip() if not (isinstance(v, float) and pd.isna(v)) else f'Col_{i}'
                     for i, v in enumerate(header_vals)]
    except Exception:
        col_names = [f'Col_{i}' for i in range(len(raw_df.columns))]

    # Data rows = everything after the header row
    data_df = raw_df.iloc[header_row_idx + 1:].copy().reset_index(drop=True)
    data_df.columns = col_names

    # Capture closing balance from ALL data rows (before any filtering)
    closing_val = _get_closing_from_raw(data_df)

    return data_df, closing_val


def _map_columns(df):
    """
    Map raw column names to standardised internal names
    (doc_date, doc_no, doc_type, particulars, debit, credit, closing).
    Works for any ERP export format — Tally, SAP, Oracle, Zoho, QuickBooks, custom.
    """
    col_map = {}
    already_mapped = set()   # prevent mapping two cols to same target

    def try_map(target, col_name):
        if target not in already_mapped:
            col_map[col_name] = target
            already_mapped.add(target)

    for c in df.columns:
        cl = str(c).lower().strip()

        # Date column
        if 'date' in cl and 'doc_date' not in already_mapped:
            try_map('doc_date', c)

        # Doc Type — must check before doc_no to avoid conflict
        elif ('type' in cl or 'nature' in cl) and 'doc_type' not in already_mapped \
                and 'date' not in cl:
            try_map('doc_type', c)

        # Doc No — broad match: no, num, number, ref, voucher, vch, details, doc no
        elif 'doc_no' not in already_mapped and 'date' not in cl and 'type' not in cl \
                and any(k in cl for k in ['no.','no ','num','voucher','vch','ref','detail',
                                           'invoice no','bill no']):
            try_map('doc_no', c)

        # Particulars / Narration / Description
        elif any(k in cl for k in ['particular','narration','description','remarks','remark']) \
                and 'particulars' not in already_mapped:
            try_map('particulars', c)

        # Opening balance — skip (don't map to closing)
        elif 'opening' in cl:
            pass

        # Debit — check it's not "credit" too
        elif 'debit' in cl and 'credit' not in cl and 'debit' not in already_mapped:
            try_map('debit', c)

        # Credit
        elif 'credit' in cl and 'debit' not in cl and 'credit' not in already_mapped:
            try_map('credit', c)

        # Closing / Balance (not opening)
        elif ('closing' in cl or 'balance' in cl) and 'opening' not in cl \
                and 'closing' not in already_mapped:
            try_map('closing', c)

    return col_map


def _load_any_ledger(file, is_vendor=True):
    """
    Universal ledger loader — works with ANY Excel format from any ERP.
    Reads raw, detects header, maps columns, handles all edge cases safely.
    """
    # Read all as strings first to avoid type-inference errors
    try:
        raw = pd.read_excel(file, header=None, dtype=str)
    except Exception:
        raw = pd.read_excel(file, header=None)

    # Detect header row
    header_idx = _detect_header_row(raw)

    # Build column names safely
    try:
        hdr_series = raw.iloc[header_idx]
        col_names = []
        for i, v in enumerate(hdr_series):
            try:
                cell = str(v).strip() if v is not None and not (isinstance(v, float) and pd.isna(v)) else ''
                col_names.append(cell if cell else f'_Col{i}')
            except Exception:
                col_names.append(f'_Col{i}')
    except Exception:
        col_names = [f'_Col{i}' for i in range(len(raw.columns))]

    # Data = rows after header
    data = raw.iloc[header_idx + 1:].copy().reset_index(drop=True)
    data.columns = col_names

    # Capture closing balance BEFORE filtering (closing row has no doc no)
    closing_val = _get_closing_from_raw(data)

    # Map columns
    col_map = _map_columns(data)
    df = data.rename(columns=col_map)

    # Ensure all needed columns exist
    needed = ['doc_date','doc_no','doc_type','debit','credit','closing']
    if is_vendor:
        needed += ['particulars']
    for col in needed:
        if col not in df.columns:
            df[col] = ''

    # Safe conversions — each wrapped so one bad column cannot crash the whole load
    try:
        df['doc_date'] = pd.to_datetime(df['doc_date'], errors='coerce', dayfirst=True)
    except Exception:
        df['doc_date'] = pd.NaT

    for num_col, fill in [('debit', 0), ('credit', 0)]:
        try:
            df[num_col] = pd.to_numeric(df[num_col], errors='coerce').fillna(fill)
        except Exception:
            df[num_col] = fill

    try:
        df['closing'] = pd.to_numeric(df['closing'], errors='coerce')
    except Exception:
        df['closing'] = np.nan

    # doc_no: safe string conversion
    try:
        df['doc_no'] = df['doc_no'].fillna('').astype(str).str.strip()
        df['doc_no'] = df['doc_no'].replace({'nan': '', 'None': '', 'NaN': ''})
    except Exception:
        df['doc_no'] = ''

    df['doc_no_clean'] = df['doc_no'].apply(clean_doc_number)
    df['period']       = df['doc_date'].apply(get_period)

    if is_vendor:
        try:
            df['particulars'] = df['particulars'].fillna('').astype(str)
        except Exception:
            df['particulars'] = ''
        df['particulars_ref'] = df['particulars'].apply(extract_ref_from_particulars)

    # Filter to rows that have a doc number
    df = df[df['doc_no_clean'].astype(str) != ''].reset_index(drop=True)
    df['_idx']       = df.index
    df['_remark']    = ''
    df['_match_ref'] = ''

    return df, closing_val


def load_vendor_ledger(file):
    df, closing = _load_any_ledger(file, is_vendor=True)
    return df, closing


def load_customer_ledger(file):
    df, closing = _load_any_ledger(file, is_vendor=False)
    return df, closing

# ─────────────────────────────────────────────
# RECONCILIATION ENGINE
# ─────────────────────────────────────────────

def run_reconciliation(vl_orig, cl_orig, tolerance=1.0):
    """
    STEP 1 — Identify ALL reversal rows in VL (Complete Reversal, Saleable Return, etc.)
             Extract original invoice no. from Particulars column.
             Match original invoice in VL by doc no AND validate amount is same/close.
             Sub-cases:
               (A) VL reversal ↔ VL original found (amount matches) + original ALSO in CL
                   → Remark: 'Invoice Reversed in VL | Present in CL - Needs Review'
               (B) VL reversal ↔ VL original found (amount matches) + NOT in CL
                   → Remark: 'Invoice Reversed in VL | Not in CL'
               (C) VL reversal found but NO original in VL (or amount mismatch)
                   → Remark: 'Reversal Entry - Original Not Found / Amount Mismatch'
    STEP 2 — Match remaining VL invoices vs CL invoices by doc number
    STEP 3 — Match debit notes by doc number → period+amount
    STEP 4 — Match collections by UTR → period+amount
    STEP 5 — All remaining → Unmatched
    """
    results = {
        'invoice_matched': [],
        'invoice_unmatched_vl': [],      # Tax Invoice / Sales Invoice only
        'invoice_unmatched_cl': [],
        'cn_unmatched_vl': [],           # Credit Notes unmatched (Saleable Return, Non-Saleable, Credit Note etc.)
        'dn_matched': [],
        'dn_unmatched_vl': [],
        'dn_unmatched_cl': [],
        'collection_matched': [],
        'collection_unmatched_vl': [],
        'collection_unmatched_cl': [],
        'reversal_vl_internal': [],
        'reversal_cross_ledger': [],
        'reversal_unmatched': [],
    }

    vl = vl_orig.copy()
    cl = cl_orig.copy()
    vl['_matched'] = False
    cl['_matched'] = False

    # ════════════════════════════════════════════════════
    # STEP 1: Process ALL VL Reversal entries
    # Detect by doc_type keywords — broad detection
    # ════════════════════════════════════════════════════
    vl_reversals = vl[vl['doc_type'].apply(is_reversal_type)].copy()

    # Pool of VL invoices available for matching (non-reversal, non-DN, non-collection)
    def get_vl_invoice_pool(vl_df):
        return vl_df[
            (~vl_df['_matched']) &
            (~vl_df['doc_type'].apply(is_reversal_type)) &
            (~vl_df['doc_type'].apply(is_debit_note)) &
            (~vl_df['doc_type'].apply(lambda x: is_collection(x)))
        ]

    for idx, rev_row in vl_reversals.iterrows():
        ref_particulars = rev_row.get('particulars_ref', '')   # cleaned ref from Particulars
        raw_particulars = str(rev_row.get('particulars', ''))  # raw Particulars text
        rev_doc_no      = rev_row.get('doc_no_clean', '')
        rev_amount      = round_amount(rev_row.get('debit', 0) + rev_row.get('credit', 0))
        rev_credit      = round_amount(rev_row.get('credit', 0))
        rev_debit       = round_amount(rev_row.get('debit', 0))

        orig_pool = get_vl_invoice_pool(vl)
        orig_match     = None
        match_basis_rev = ''

        # ── Method 1: Exact match on cleaned particulars ref ──
        if ref_particulars:
            m = orig_pool[orig_pool['doc_no_clean'] == ref_particulars]
            if not m.empty:
                orig_match = m.iloc[0]
                match_basis_rev = 'Particulars Reference (Exact)'

        # ── Method 2: Scan all words in raw Particulars against VL doc nos ──
        if orig_match is None:
            words = re.findall(r'[A-Z0-9]{4,}', raw_particulars.upper())
            for word in words:
                m = orig_pool[orig_pool['doc_no_clean'] == word]
                if not m.empty:
                    orig_match = m.iloc[0]
                    match_basis_rev = f'Particulars Word Match ({word})'
                    break

        # ── Method 3: Partial substring match (first 8 chars of ref) ──
        if orig_match is None and ref_particulars and len(ref_particulars) >= 5:
            prefix = ref_particulars[:8]
            m = orig_pool[orig_pool['doc_no_clean'].str.startswith(prefix, na=False)]
            if not m.empty:
                orig_match = m.iloc[0]
                match_basis_rev = 'Particulars Partial Match'

        # ── Method 4: Same period + same amount (fallback) ──
        if orig_match is None and rev_amount > 0:
            m = orig_pool[
                (orig_pool['period'] == rev_row.get('period', '')) &
                (abs(orig_pool['debit'] + orig_pool['credit'] - rev_amount) <= tolerance)
            ]
            if not m.empty:
                orig_match = m.iloc[0]
                match_basis_rev = 'Period + Amount Match'

        # ── Amount Validation: reversal amount must be same/close to original ──
        # STRICT CHECK: only compare total amounts — not cross debit/credit
        # This prevents false matches where amounts are completely different
        amount_valid = False
        if orig_match is not None:
            orig_amount = round_amount(
                orig_match.get('debit', 0) + orig_match.get('credit', 0)
            )
            # Only accept if total amounts are within tolerance
            if orig_amount > 0 and abs(rev_amount - orig_amount) <= tolerance:
                amount_valid = True

        if orig_match is not None and amount_valid:
            orig_vl_idx = orig_match['_idx']
            orig_doc_no = str(orig_match.get('doc_no', ''))
            orig_amount_val = round_amount(orig_match.get('debit', 0) + orig_match.get('credit', 0))

            # Check if original invoice also exists in CL
            cl_pool = cl[
                (~cl['_matched']) &
                (~cl['doc_type'].apply(is_debit_note)) &
                (~cl['doc_type'].apply(lambda x: is_collection(x, '')))
            ]
            cl_for_orig = cl_pool[cl_pool['doc_no_clean'] == orig_match['doc_no_clean']]

            # Mark VL reversal row
            vl.at[idx, '_matched'] = True
            vl.at[idx, '_match_ref'] = orig_doc_no

            # Mark VL original row
            vl.at[orig_vl_idx, '_matched'] = True
            vl.at[orig_vl_idx, '_match_ref'] = str(rev_row.get('doc_no', ''))

            if not cl_for_orig.empty:
                # ── CASE A: Invoice reversed in VL but ALSO present in CL ──
                cl_row = cl_for_orig.iloc[0]
                cl.at[cl_row['_idx'], '_matched'] = True
                cl.at[cl_row['_idx'], '_remark'] = 'Invoice Reversed in VL — Needs Review'
                cl.at[cl_row['_idx'], '_match_ref'] = orig_doc_no

                vl.at[idx, '_remark'] = 'Reversal Entry — Invoice Also in CL'
                vl.at[orig_vl_idx, '_remark'] = 'Invoice Reversed in VL — Also in CL'

                results['reversal_cross_ledger'].append({
                    'VL Original Doc No': orig_doc_no,
                    'VL Original Date': orig_match.get('doc_date', ''),
                    'VL Original Type': orig_match.get('doc_type', ''),
                    'VL Original Debit': orig_match.get('debit', 0),
                    'VL Original Credit': orig_match.get('credit', 0),
                    'VL Reversal Doc No': str(rev_row.get('doc_no', '')),
                    'VL Reversal Date': rev_row.get('doc_date', ''),
                    'VL Reversal Type': rev_row.get('doc_type', ''),
                    'VL Reversal Debit': rev_row.get('debit', 0),
                    'VL Reversal Credit': rev_row.get('credit', 0),
                    'VL Reversal Particulars': raw_particulars,
                    'CL Doc No': cl_row.get('doc_no', ''),
                    'CL Date': cl_row.get('doc_date', ''),
                    'CL Type': cl_row.get('doc_type', ''),
                    'CL Debit': cl_row.get('debit', 0),
                    'CL Credit': cl_row.get('credit', 0),
                    'Match Basis': match_basis_rev,
                    'Amount Match': f'VL={rev_amount} | Orig={orig_amount_val}',
                    'Remark': 'Invoice Reversed in VL but Present in CL — Needs Review',
                })
            else:
                # ── CASE B: Pure VL internal reversal — not in CL ──
                vl.at[idx, '_remark'] = 'Reversal Entry — Not in CL'
                vl.at[orig_vl_idx, '_remark'] = 'Invoice Reversed in VL — Not in CL'

                results['reversal_vl_internal'].append({
                    'VL Original Doc No': orig_doc_no,
                    'VL Original Date': orig_match.get('doc_date', ''),
                    'VL Original Type': orig_match.get('doc_type', ''),
                    'VL Original Debit': orig_match.get('debit', 0),
                    'VL Original Credit': orig_match.get('credit', 0),
                    'VL Reversal Doc No': str(rev_row.get('doc_no', '')),
                    'VL Reversal Date': rev_row.get('doc_date', ''),
                    'VL Reversal Type': rev_row.get('doc_type', ''),
                    'VL Reversal Debit': rev_row.get('debit', 0),
                    'VL Reversal Credit': rev_row.get('credit', 0),
                    'VL Reversal Particulars': raw_particulars,
                    'Match Basis': match_basis_rev,
                    'Amount Match': f'VL={rev_amount} | Orig={orig_amount_val}',
                    'Remark': 'Invoice Reversed in VL — Not Present in CL',
                })
        else:
            # ── CASE C: No original found OR amount mismatch ──
            # NOTE: Case C rows are NOT marked as matched — they flow to Steps 2/2B
            is_amount_mismatch = (orig_match is not None and not amount_valid)

            # Remark in VL ledger — clean, no amounts embedded
            if is_amount_mismatch:
                vl_remark = 'Unmatched — Amount Mismatch'
            else:
                vl_remark = 'Unmatched — Original Invoice Not Found in VL'

            # Separate amount columns for the annexure report
            rev_amt_col  = rev_amount
            orig_amt_col = round_amount(orig_match.get('debit', 0) + orig_match.get('credit', 0)) if orig_match is not None else 0
            reason_col   = 'Amount Mismatch' if is_amount_mismatch else 'Original Invoice Not Found in VL'

            vl.at[idx, '_remark']  = vl_remark
            # Do NOT set _matched=True — Case C continues to Step 2/2B matching

            results['reversal_unmatched'].append({
                'VL Doc No':          str(rev_row.get('doc_no', '')),
                'VL Date':            rev_row.get('doc_date', ''),
                'VL Type':            rev_row.get('doc_type', ''),
                'Particulars':        raw_particulars,
                'VL Reversal Amount': rev_amt_col,
                'Original Amount':    orig_amt_col,
                'VL Debit':           rev_row.get('debit', 0),
                'VL Credit':          rev_row.get('credit', 0),
                'Reason':             reason_col,
                'Remark':             vl_remark,
            })

    # ════════════════════════════════════════════════════
    # STEP 2: Match Invoices by Document Number (VL vs CL)
    # Excludes Credit Notes (handled separately in Step 2B)
    # ════════════════════════════════════════════════════
    vl_inv = vl[
        (~vl['_matched']) &
        (~vl['doc_type'].apply(is_debit_note)) &
        (~vl['doc_type'].apply(lambda x: is_collection(x, ''))) &
        (~vl['doc_type'].apply(is_reversal_type)) &
        (~vl['doc_type'].apply(is_credit_note))   # Credit Notes handled in Step 2B
    ].copy()

    cl_inv = cl[
        (~cl['_matched']) &
        (~cl['doc_type'].apply(is_debit_note)) &
        (~cl['doc_type'].apply(lambda x: is_collection(x, ''))) &
        (~cl['doc_type'].apply(lambda x: is_discount_or_prn(x)))  # Discount/PRN handled in Step 2B
    ].copy()

    for idx, vrow in vl_inv.iterrows():
        matches = cl_inv[(cl_inv['doc_no_clean'] == vrow['doc_no_clean']) & (~cl_inv['_matched'])]
        if not matches.empty:
            crow = matches.iloc[0]
            vl.at[idx, '_matched'] = True
            vl.at[idx, '_remark'] = 'Matched — Invoice'
            vl.at[idx, '_match_ref'] = str(crow.get('doc_no', ''))
            cl.at[crow['_idx'], '_matched'] = True
            cl.at[crow['_idx'], '_remark'] = 'Matched — Invoice'
            cl.at[crow['_idx'], '_match_ref'] = str(vrow.get('doc_no', ''))
            cl_inv.at[crow.name, '_matched'] = True
            results['invoice_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Debit': vrow.get('debit', 0),
                'VL Credit': vrow.get('credit', 0),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Debit': crow.get('debit', 0),
                'CL Credit': crow.get('credit', 0),
                'Match Basis': 'Document Number',
                'Match Type': 'Invoice',
                'Remark': 'Matched — Invoice',
            })

    # ════════════════════════════════════════════════════
    # STEP 2B: Match VL Credit Notes against CL Discount
    #          Debit Notes and PRN entries
    # Matching order:
    #   1st — Document Number (exact)
    #   2nd — Period + Amount (within tolerance)
    # ════════════════════════════════════════════════════
    vl_cn = vl[
        (~vl['_matched']) &
        (vl['doc_type'].apply(is_credit_note))
    ].copy()

    cl_disc = cl[
        (~cl['_matched']) &
        (cl.apply(lambda r: is_discount_or_prn(r.get('doc_type', ''), r.get('doc_no', '')), axis=1))
    ].copy()

    # Also include any unmatched CL entries that could be credit note counterparts
    cl_any_unmatched = cl[
        (~cl['_matched']) &
        (~cl['doc_type'].apply(lambda x: is_collection(x, '')))
    ].copy()

    for idx, vrow in vl_cn.iterrows():
        matched = False
        basis = ''
        crow = None

        # 1st: match by document number against discount/PRN pool
        if not cl_disc.empty:
            doc_m = cl_disc[(cl_disc['doc_no_clean'] == vrow['doc_no_clean']) & (~cl_disc['_matched'])]
            if not doc_m.empty:
                crow = doc_m.iloc[0]
                matched = True
                basis = 'Document Number (Credit Note ↔ Discount/PRN)'

        # 2nd: match by period + amount against discount/PRN pool
        if not matched and not cl_disc.empty:
            vamt = round_amount(vrow['debit'] + vrow['credit'])
            if vamt > 0:
                amt_m = cl_disc[
                    (cl_disc['period'] == vrow['period']) &
                    (abs(cl_disc['debit'] + cl_disc['credit'] - vamt) <= tolerance) &
                    (~cl_disc['_matched'])
                ]
                if not amt_m.empty:
                    crow = amt_m.iloc[0]
                    matched = True
                    basis = 'Period + Amount (Credit Note ↔ Discount/PRN)'

        # 3rd: try doc number match against ALL unmatched CL (broader fallback)
        if not matched:
            doc_m2 = cl_any_unmatched[
                (cl_any_unmatched['doc_no_clean'] == vrow['doc_no_clean']) &
                (~cl_any_unmatched['_matched'])
            ]
            if not doc_m2.empty:
                crow = doc_m2.iloc[0]
                matched = True
                basis = 'Document Number (Credit Note ↔ CL Entry)'

        if matched and crow is not None:
            vl.at[idx, '_matched'] = True
            vl.at[idx, '_remark'] = f'Matched — Credit Note ({basis.split("(")[0].strip()})'
            vl.at[idx, '_match_ref'] = str(crow.get('doc_no', ''))
            cl.at[crow['_idx'], '_matched'] = True
            cl.at[crow['_idx'], '_remark'] = f'Matched — Credit Note / Discount-PRN ({basis.split("(")[0].strip()})'
            cl.at[crow['_idx'], '_match_ref'] = str(vrow.get('doc_no', ''))
            if crow.name in cl_disc.index:
                cl_disc.at[crow.name, '_matched'] = True
            if crow.name in cl_any_unmatched.index:
                cl_any_unmatched.at[crow.name, '_matched'] = True
            results['dn_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Debit': vrow.get('debit', 0),
                'VL Credit': vrow.get('credit', 0),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Debit': crow.get('debit', 0),
                'CL Credit': crow.get('credit', 0),
                'Match Basis': basis,
                'Match Type': 'Credit Note ↔ Discount / PRN',
                'Remark': f'Matched — Credit Note vs Discount/PRN ({basis.split("(")[0].strip()})',
            })

    # ════════════════════════════════════════════════════
    # STEP 3: Match Debit Notes
    # ════════════════════════════════════════════════════
    vl_dn = vl[(~vl['_matched']) & (vl['doc_type'].apply(is_debit_note))].copy()
    cl_dn = cl[(~cl['_matched']) & (cl['doc_type'].apply(is_debit_note))].copy()

    for idx, vrow in vl_dn.iterrows():
        matched = False
        basis = ''
        doc_matches = cl_dn[(cl_dn['doc_no_clean'] == vrow['doc_no_clean']) & (~cl_dn['_matched'])]
        if not doc_matches.empty:
            crow = doc_matches.iloc[0]
            matched = True
            basis = 'Document Number'
        else:
            vamt = round_amount(vrow['debit'] + vrow['credit'])
            period_matches = cl_dn[
                (cl_dn['period'] == vrow['period']) &
                (abs(cl_dn['debit'] + cl_dn['credit'] - vamt) <= tolerance) &
                (~cl_dn['_matched'])
            ]
            if not period_matches.empty:
                crow = period_matches.iloc[0]
                matched = True
                basis = 'Period + Amount'

        if matched:
            vl.at[idx, '_matched'] = True
            vl.at[idx, '_remark'] = f'Matched — Debit Note ({basis})'
            vl.at[idx, '_match_ref'] = str(crow.get('doc_no', ''))
            cl.at[crow['_idx'], '_matched'] = True
            cl.at[crow['_idx'], '_remark'] = f'Matched — Debit Note ({basis})'
            cl.at[crow['_idx'], '_match_ref'] = str(vrow.get('doc_no', ''))
            cl_dn.at[crow.name, '_matched'] = True
            results['dn_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Debit': vrow.get('debit', 0),
                'VL Credit': vrow.get('credit', 0),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Debit': crow.get('debit', 0),
                'CL Credit': crow.get('credit', 0),
                'Match Basis': basis,
                'Match Type': 'Debit Note',
                'Remark': f'Matched - Debit Note ({basis})',
            })

    # ════════════════════════════════════════════════════
    # STEP 4: Match Collections by UTR or Period+Amount
    # ════════════════════════════════════════════════════
    vl_col = vl[(~vl['_matched']) & (vl['doc_type'].apply(lambda x: is_collection(x, '')))].copy()
    cl_col = cl[(~cl['_matched']) & (cl['doc_type'].apply(lambda x: is_collection(x, '')))].copy()

    vl_col['utr'] = vl_col.apply(lambda r: extract_utr(str(r.get('particulars', '')) + ' ' + str(r.get('doc_no', ''))), axis=1)
    cl_col['utr'] = cl_col.apply(lambda r: extract_utr(str(r.get('doc_no', '')) + ' ' + str(r.get('doc_type', ''))), axis=1)

    for idx, vrow in vl_col.iterrows():
        matched = False
        basis = ''
        if vrow['utr']:
            utr_matches = cl_col[(cl_col['utr'] == vrow['utr']) & (cl_col['utr'] != '') & (~cl_col['_matched'])]
            if not utr_matches.empty:
                crow = utr_matches.iloc[0]
                matched = True
                basis = 'UTR Number'

        if not matched:
            vamt = round_amount(vrow['debit'] + vrow['credit'])
            amt_matches = cl_col[
                (cl_col['period'] == vrow['period']) &
                (abs(cl_col['debit'] + cl_col['credit'] - vamt) <= tolerance) &
                (~cl_col['_matched'])
            ]
            if not amt_matches.empty:
                crow = amt_matches.iloc[0]
                matched = True
                basis = 'Period + Amount'

        if matched:
            vl.at[idx, '_matched'] = True
            vl.at[idx, '_remark'] = f'Matched — Collection ({basis})'
            vl.at[idx, '_match_ref'] = str(crow.get('doc_no', ''))
            cl.at[crow['_idx'], '_matched'] = True
            cl.at[crow['_idx'], '_remark'] = f'Matched — Collection ({basis})'
            cl.at[crow['_idx'], '_match_ref'] = str(vrow.get('doc_no', ''))
            cl_col.at[crow.name, '_matched'] = True
            results['collection_matched'].append({
                'VL Doc No': vrow.get('doc_no', ''),
                'VL Date': vrow.get('doc_date', ''),
                'VL Type': vrow.get('doc_type', ''),
                'VL Amount': vrow.get('debit', 0) + vrow.get('credit', 0),
                'VL UTR': vrow.get('utr', ''),
                'CL Doc No': crow.get('doc_no', ''),
                'CL Date': crow.get('doc_date', ''),
                'CL Type': crow.get('doc_type', ''),
                'CL Amount': crow.get('debit', 0) + crow.get('credit', 0),
                'CL UTR': crow.get('utr', ''),
                'Match Basis': basis,
                'Match Type': 'Collection',
                'Remark': f'Matched - Collection ({basis})',
            })

    # ════════════════════════════════════════════════════
    # STEP 5: Mark all remaining rows as Unmatched
    # Credit Notes go to cn_unmatched_vl (separate from invoices)
    # ════════════════════════════════════════════════════
    for idx, r in vl[~vl['_matched']].iterrows():
        doc_t = str(r.get('doc_type', ''))
        if is_credit_note(doc_t):
            unmatched_remark = 'Unmatched — Credit Note'
        elif is_reversal_type(doc_t):
            unmatched_remark = 'Unmatched — Reversal Entry'
        else:
            unmatched_remark = 'Unmatched — Invoice'
        vl.at[idx, '_remark'] = unmatched_remark
        entry = {
            'Doc No': r.get('doc_no', ''),
            'Date': r.get('doc_date', ''),
            'Type': r.get('doc_type', ''),
            'Particulars': r.get('particulars', ''),
            'Debit': r.get('debit', 0),
            'Credit': r.get('credit', 0),
            'Source': 'Vendor Ledger',
            'Remark': unmatched_remark,
        }
        if is_debit_note(r.get('doc_type', '')):
            results['dn_unmatched_vl'].append(entry)
        elif is_collection(r.get('doc_type', '')):
            results['collection_unmatched_vl'].append(entry)
        elif is_credit_note(r.get('doc_type', '')):
            # Credit Notes (Saleable Return, Non-Saleable, Credit Note, etc.) — separate bucket
            results['cn_unmatched_vl'].append(entry)
        else:
            # Only pure invoices / tax invoices / sales invoices
            results['invoice_unmatched_vl'].append(entry)

    for idx, r in cl[~cl['_matched']].iterrows():
        doc_t = str(r.get('doc_type', ''))
        if is_credit_note(doc_t):
            cl_unmatched_remark = 'Unmatched — Credit Note'
        else:
            cl_unmatched_remark = 'Unmatched — Invoice'
        cl.at[idx, '_remark'] = cl_unmatched_remark
        entry = {
            'Doc No': r.get('doc_no', ''),
            'Date': r.get('doc_date', ''),
            'Type': r.get('doc_type', ''),
            'Debit': r.get('debit', 0),
            'Credit': r.get('credit', 0),
            'Source': 'Customer Ledger',
            'Remark': cl_unmatched_remark,
        }
        if is_debit_note(r.get('doc_type', '')):
            results['dn_unmatched_cl'].append(entry)
        elif is_collection(r.get('doc_type', '')):
            results['collection_unmatched_cl'].append(entry)
        else:
            results['invoice_unmatched_cl'].append(entry)

    results['vl_annotated'] = vl
    results['cl_annotated'] = cl

    return results

# ─────────────────────────────────────────────
# EXCEL EXPORT — with Vendor & Customer Ledger tabs + Remarks
# ─────────────────────────────────────────────

def build_excel(results, vl_orig, cl_orig, VL='Vendor', CL='Customer'):
    output = BytesIO()
    wb = openpyxl.Workbook()

    VL_COLOR = '1A3A6B'; CL_COLOR = '1A6B45'
    VL_LIGHT = 'D6E4FF'; CL_LIGHT = 'D6F5EA'
    MTH_COLOR = '1A6B45'; UNM_COLOR = 'A32035'
    MTH_FILL = 'C6EFCE'; UNM_FILL = 'FFC7CE'; REV_FILL = 'FFEB9C'; DARK = '1C2130'
    COLORS = {'matched_fill': MTH_FILL, 'unmatched_fill': UNM_FILL, 'reversal_fill': REV_FILL, 'border': 'C0C8D8'}

    thin = Side(style='thin', color=COLORS['border'])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def mk_fill(h): return PatternFill(fill_type='solid', fgColor=h)

    def style_header(ws, headers, row=1, color=DARK):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            cell.fill = mk_fill(color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 30

    def auto_width(ws, min_w=10, max_w=48):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try: max_len = max(max_len, len(str(cell.value or '')))
                except: pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, max_len + 2))

    def fmt_date(val):
        try:
            if pd.isna(val): return ''
            return pd.to_datetime(val).strftime('%d-%b-%Y')
        except: return str(val)

    def write_cell(ws, row, col, val):
        cell = ws.cell(row=row, column=col)
        if isinstance(val, (pd.Timestamp, datetime)): cell.value = fmt_date(val)
        elif isinstance(val, float) and not pd.isna(val):
            cell.value = round(val, 2); cell.number_format = '#,##0.00'
        elif not isinstance(val, str) and pd.isna(val): cell.value = ''
        else: cell.value = val
        return cell

    def ssum(lst, key):
        try: return sum(float(d.get(key, 0) or 0) for d in lst)
        except: return 0.0

    def pct(m, t): return f'{round(m/t*100,1)}%' if t else '0%'

    cn_list = [r for r in results['dn_matched'] if 'Credit Note' in str(r.get('Match Type',''))]
    dn_list = [r for r in results['dn_matched'] if 'Credit Note' not in str(r.get('Match Type',''))]

    inv_m_c = len(results['invoice_matched']); inv_m_v = ssum(results['invoice_matched'],'VL Debit')+ssum(results['invoice_matched'],'VL Credit')
    inv_uvl_c = len(results['invoice_unmatched_vl']); inv_uvl_v = ssum(results['invoice_unmatched_vl'],'Debit')+ssum(results['invoice_unmatched_vl'],'Credit')
    inv_ucl_c = len(results['invoice_unmatched_cl']); inv_ucl_v = ssum(results['invoice_unmatched_cl'],'Debit')+ssum(results['invoice_unmatched_cl'],'Credit')
    dn_m_c = len(dn_list); dn_m_v = ssum(dn_list,'VL Debit')+ssum(dn_list,'VL Credit')
    dn_uvl_c = len(results['dn_unmatched_vl']); dn_uvl_v = ssum(results['dn_unmatched_vl'],'Debit')+ssum(results['dn_unmatched_vl'],'Credit')
    dn_ucl_c = len(results['dn_unmatched_cl']); dn_ucl_v = ssum(results['dn_unmatched_cl'],'Debit')+ssum(results['dn_unmatched_cl'],'Credit')
    cn_m_c = len(cn_list); cn_m_v = ssum(cn_list,'VL Debit')+ssum(cn_list,'VL Credit')
    col_m_c = len(results['collection_matched']); col_m_v = ssum(results['collection_matched'],'VL Amount')
    col_uvl_c = len(results['collection_unmatched_vl']); col_uvl_v = ssum(results['collection_unmatched_vl'],'Debit')+ssum(results['collection_unmatched_vl'],'Credit')
    col_ucl_c = len(results['collection_unmatched_cl']); col_ucl_v = ssum(results['collection_unmatched_cl'],'Debit')+ssum(results['collection_unmatched_cl'],'Credit')
    rcl_c = len(results['reversal_cross_ledger']); rcl_v = ssum(results['reversal_cross_ledger'],'VL Original Debit')+ssum(results['reversal_cross_ledger'],'VL Original Credit')
    rvl_c = len(results['reversal_vl_internal']); rvl_v = ssum(results['reversal_vl_internal'],'VL Original Debit')+ssum(results['reversal_vl_internal'],'VL Original Credit')
    run_c = len(results['reversal_unmatched']); run_v = ssum(results['reversal_unmatched'],'VL Debit')+ssum(results['reversal_unmatched'],'VL Credit')
    mis_c = len([r for r in results['reversal_unmatched'] if r.get('Reason','')=='Amount Mismatch'])
    miss_c = run_c - mis_c
    inv_vl_t=inv_m_c+inv_uvl_c; inv_cl_t=inv_m_c+inv_ucl_c; inv_vl_v=inv_m_v+inv_uvl_v; inv_cl_v=inv_m_v+inv_ucl_v
    dn_vl_t=dn_m_c+dn_uvl_c; dn_cl_t=dn_m_c+dn_ucl_c; dn_vl_v=dn_m_v+dn_uvl_v; dn_cl_v=dn_m_v+dn_ucl_v
    col_vl_t=col_m_c+col_uvl_c; col_cl_t=col_m_c+col_ucl_c; col_vl_v=col_m_v+col_uvl_v; col_cl_v=col_m_v+col_ucl_v
    tot_m_c=inv_m_c+dn_m_c+cn_m_c+col_m_c; tot_m_v=inv_m_v+dn_m_v+cn_m_v+col_m_v

    # ══ SUMMARY SHEET ══
    ws_sum = wb.active; ws_sum.title = 'Summary'; ws_sum.sheet_view.showGridLines = False

    ws_sum.merge_cells('A1:M1')
    tc = ws_sum['A1']; tc.value = f'⚖️  LEDGER RECONCILIATION — {VL}  vs  {CL}'
    tc.font = Font(bold=True, size=14, color='FFFFFF', name='Calibri')
    tc.fill = mk_fill(DARK); tc.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 38

    ws_sum.merge_cells('A2:M2')
    ws_sum['A2'].value = f'Generated: {datetime.now().strftime("%d-%b-%Y %H:%M")}   |   Blue = {VL}   |   Teal = {CL}   |   Subtotal row is at top (row 5)'
    ws_sum['A2'].font = Font(italic=True, size=9, color='888888', name='Calibri')
    ws_sum.row_dimensions[2].height = 16

    # Group header row 3
    grp3 = [('A3','A3','Category',DARK),('B3','E3',f'{VL} — Vendor Ledger',VL_COLOR),
             ('F3','I3',f'{CL} — Customer Ledger',CL_COLOR),
             ('J3','J3','Match %',DARK),('K3','K3','Remarks',DARK),('L3','M3','Annexure',DARK)]
    for s,e,lbl,bg in grp3:
        if s != e: ws_sum.merge_cells(f'{s}:{e}')
        c = ws_sum[s]; c.value = lbl
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        c.fill = mk_fill(bg); c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
    ws_sum.row_dimensions[3].height = 24

    # Col sub-headers row 4
    ch4 = [(VL+' Total\nCount',VL_COLOR),(VL+' Total\nValue',VL_COLOR),
           ('Matched\nCount',VL_COLOR),('Matched\nValue',VL_COLOR),
           (CL+' Total\nCount',CL_COLOR),(CL+' Total\nValue',CL_COLOR),
           ('Unmatched\nCount',CL_COLOR),('Unmatched\nValue',CL_COLOR)]
    for c_idx, lbl in enumerate(['Category']+[h for h,_ in ch4]+['Match %','Remarks','Annexure Sheet','Mis / Miss'],1):
        bg = ch4[c_idx-2][1] if 2<=c_idx<=9 else DARK
        cell = ws_sum.cell(row=4, column=c_idx, value=lbl)
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
        cell.fill = mk_fill(bg); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = border
    ws_sum.row_dimensions[4].height = 36

    # SUBTOTAL row 5 — formulas filled after data written
    for c_idx in range(1,14):
        cell = ws_sum.cell(row=5, column=c_idx)
        cell.fill = mk_fill('2A3A5A'); cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10); cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.cell(row=5, column=1).value = 'SUBTOTAL ▲'
    ws_sum.row_dimensions[5].height = 24

    DS = 6  # data start row
    rows_data = [
        (f'Invoices', inv_vl_t, inv_vl_v, inv_m_c, inv_m_v, inv_cl_t, inv_cl_v, inv_ucl_c, inv_ucl_v, pct(inv_m_c,inv_vl_t), 'Matched by Document Number', 'Inv-Matched', '', 'EEF3FF'),
        (f'Credit Notes ↔ Disc DN/PRN', cn_m_c, cn_m_v, cn_m_c, cn_m_v, '-', 0, '-', 0, pct(cn_m_c,max(cn_m_c,1)), f'{VL} Credit Note vs {CL} Discount/PRN', 'DN-CN-Matched', '', 'EEF3FF'),
        (f'Debit Notes', dn_vl_t, dn_vl_v, dn_m_c, dn_m_v, dn_cl_t, dn_cl_v, dn_ucl_c, dn_ucl_v, pct(dn_m_c,dn_vl_t), 'Matched by Doc No / Period+Amount', 'DN-CN-Matched', '', 'EEF8F3'),
        (f'Collections', col_vl_t, col_vl_v, col_m_c, col_m_v, col_cl_t, col_cl_v, col_ucl_c, col_ucl_v, pct(col_m_c,col_vl_t), 'Matched by UTR / Period+Amount', 'Coll-Matched', '', 'EEF8F3'),
        (f'Reversal — Also in {CL} (A)', rcl_c, rcl_v, rcl_c, rcl_v, rcl_c, rcl_v, 0, 0, '⚠️ Review', f'Reversed in {VL} but in {CL}', 'AnnexA-CrossLedger', '', 'FFF8EC'),
        (f'Reversal — Not in {CL} (B)', rvl_c*2, rvl_v, rvl_c, rvl_v, '-', 0, '-', 0, '✅ 100%', f'Reversed in {VL} only', 'AnnexB-VL-Internal', '', 'EEF3FF'),
        (f'Reversal — Amt Mismatch (C1)', mis_c, 0, 0, 0, '-', 0, mis_c, 0, '❌ 0%', 'Amount does not match original', 'AnnexC1-AmtMismatch', f'Mismatch: {mis_c}', 'FFE8E8'),
        (f'Reversal — Missing Orig (C2)', miss_c, 0, 0, 0, '-', 0, miss_c, 0, '❌ 0%', 'Original invoice not found in VL', 'AnnexC2-MissingOrig', f'Missing: {miss_c}', 'FFE8E8'),
    ]

    for off, rd in enumerate(rows_data):
        lbl,vl_c,vl_v,mc,mv,cl_c,cl_v,uc,uv,p,rem,ann,mm,rf = rd
        r = DS + off
        cols_def = [(1,lbl,DARK,rf,False),(2,vl_c,VL_COLOR,VL_LIGHT,False),(3,vl_v,VL_COLOR,VL_LIGHT,True),
                    (4,mc,MTH_COLOR,MTH_FILL,False),(5,mv,MTH_COLOR,MTH_FILL,True),
                    (6,cl_c,CL_COLOR,CL_LIGHT,False),(7,cl_v,CL_COLOR,CL_LIGHT,True),
                    (8,uc,UNM_COLOR,UNM_FILL,False),(9,uv,UNM_COLOR,UNM_FILL,True),
                    (10,p,DARK,rf,False),(11,rem,DARK,rf,False),(12,f'→ {ann}',DARK,rf,False),(13,mm,DARK,rf,False)]
        for ci,val,fg,fill,is_val in cols_def:
            cell = ws_sum.cell(row=r, column=ci, value=val)
            cell.fill = mk_fill(fill); cell.border = border
            if ci == 1: cell.font = Font(bold=True, name='Calibri', size=10)
            elif is_val and isinstance(val,(int,float)):
                cell.font = Font(bold=True, color=fg, name='Calibri', size=10)
                cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal='right', vertical='center')
            elif ci in [2,4,6,8] and isinstance(val,(int,float)):
                cell.font = Font(bold=True, color=fg, name='Calibri', size=10)
                cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right', vertical='center')
            elif ci == 10:
                fc2 = UNM_COLOR if any(x in str(val) for x in ['❌','⚠️']) else MTH_COLOR
                cell.font = Font(bold=True, color=fc2, name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif ci == 12:
                cell.font = Font(color='1A6BCC', bold=True, underline='single', name='Calibri', size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = Font(name='Calibri', size=9)
                cell.alignment = Alignment(vertical='center', wrap_text=(ci in [1,11]))
        ws_sum.row_dimensions[r].height = 22

    DE = DS + len(rows_data) - 1

    # Fill subtotal formulas at row 5
    num_fmt_map = {2:'#,##0',3:'#,##0.00',4:'#,##0',5:'#,##0.00',6:'#,##0',7:'#,##0.00',8:'#,##0',9:'#,##0.00'}
    for ci, nfmt in num_fmt_map.items():
        cl_l = get_column_letter(ci)
        cell = ws_sum.cell(row=5, column=ci, value=f'=SUBTOTAL(9,{cl_l}{DS}:{cl_l}{DE})')
        cell.number_format = nfmt; cell.fill = mk_fill('2A3A5A')
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10); cell.border = border
        cell.alignment = Alignment(horizontal='right', vertical='center')

    # Total row
    tr = DE + 1
    tv = ['TOTAL', inv_vl_t+dn_vl_t+col_vl_t, inv_vl_v+dn_vl_v+col_vl_v,
          tot_m_c, tot_m_v, inv_cl_t+dn_cl_t+col_cl_t, inv_cl_v+dn_cl_v+col_cl_v,
          inv_ucl_c+dn_ucl_c+col_ucl_c, inv_ucl_v+dn_ucl_v+col_ucl_v,
          pct(tot_m_c,inv_vl_t+dn_vl_t+col_vl_t), '', '', '']
    for ci, val in enumerate(tv, 1):
        cell = ws_sum.cell(row=tr, column=ci, value=val)
        cell.fill = mk_fill(DARK); cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        cell.border = border; cell.alignment = Alignment(horizontal='center', vertical='center')
        if ci in [3,5,7,9] and isinstance(val, float): cell.number_format = '#,##0.00'
    ws_sum.row_dimensions[tr].height = 28

    for i,w in enumerate([42,12,16,12,16,12,16,12,16,10,40,24,18],1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════
    # SHEET 2: VENDOR LEDGER WITH REMARKS
    # ══════════════════════════════════════════
    vl_ann = vl_orig
    ws_vl = wb.create_sheet(f'{VL[:18]} Vendor Ledger'[:31])
    ws_vl.sheet_view.showGridLines = False
    ws_vl.freeze_panes = 'A4'

    vl_display_cols = ['doc_date', 'doc_no', 'doc_type', 'particulars', 'debit', 'credit', 'closing']
    vl_display_cols = [c for c in vl_display_cols if c in vl_ann.columns]
    vl_display_cols += ['_remark', '_match_ref']
    ncols_vl = len(vl_display_cols)

    # Row 1: Title
    ws_vl.merge_cells(f'A1:{get_column_letter(ncols_vl)}1')
    title_vl = ws_vl['A1']
    title_vl.value = f'📘 {VL} — VENDOR LEDGER WITH REMARKS'
    title_vl.font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
    title_vl.fill = mk_fill('1A3A6B')
    title_vl.alignment = Alignment(horizontal='center', vertical='center')
    ws_vl.row_dimensions[1].height = 28

    # Row 2: SUBTOTAL at top (formulas filled after data)
    for c_idx in range(1, ncols_vl + 1):
        cell = ws_vl.cell(row=2, column=c_idx)
        cell.fill = mk_fill('2A3A5A')
        cell.border = border
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_vl.cell(row=2, column=1).value = 'SUBTOTAL ▲'
    ws_vl.row_dimensions[2].height = 22

    # Row 3: Headers
    vl_hmap = {'doc_date':'Doc Date','doc_no':'Doc No','doc_type':'Doc Type',
               'particulars':'Particulars','debit':'Debit','credit':'Credit',
               'closing':'Closing Balance','_remark':'Remark','_match_ref':'Matched With'}
    headers = [vl_hmap.get(c, c) for c in vl_display_cols]
    style_header(ws_vl, headers, row=3, color='1A3A6B')

    debit_col_vl  = (vl_display_cols.index('debit')  + 1) if 'debit'  in vl_display_cols else None
    credit_col_vl = (vl_display_cols.index('credit') + 1) if 'credit' in vl_display_cols else None

    # Data starts at row 4
    VL_DATA_START = 4
    for r_idx, (_, row) in enumerate(vl_ann[vl_display_cols].iterrows(), VL_DATA_START):
        remark = str(row.get('_remark', ''))
        if 'Unmatched' in remark or 'Mismatch' in remark or 'Not Found' in remark:
            fill = COLORS['unmatched_fill']
        elif 'Reversal Entry' in remark or 'Invoice Reversed' in remark:
            fill = COLORS['reversal_fill']
        else:
            fill = COLORS['matched_fill']
        for c_idx, col in enumerate(vl_display_cols, 1):
            val = row[col]
            cell = write_cell(ws_vl, r_idx, c_idx, val)
            cell.fill = mk_fill(fill)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            cell.font = Font(name='Calibri', size=9)
        ws_vl.row_dimensions[r_idx].height = 18

    vl_data_end = VL_DATA_START + len(vl_ann) - 1

    # Fill SUBTOTAL formulas at row 2
    for col_i in [debit_col_vl, credit_col_vl]:
        if col_i:
            cl_l = get_column_letter(col_i)
            cell = ws_vl.cell(row=2, column=col_i,
                              value=f'=SUBTOTAL(9,{cl_l}{VL_DATA_START}:{cl_l}{vl_data_end})')
            cell.number_format = '#,##0.00'
            cell.fill = mk_fill('2A3A5A')
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')
    auto_width(ws_vl)

    # ══════════════════════════════════════════
    # SHEET 3: CUSTOMER LEDGER WITH REMARKS
    # ══════════════════════════════════════════
    cl_ann = cl_orig
    ws_cl = wb.create_sheet(f'{CL[:18]} Customer Ledger'[:31])
    ws_cl.sheet_view.showGridLines = False
    ws_cl.freeze_panes = 'A4'

    cl_display_cols = ['doc_date', 'doc_no', 'doc_type', 'debit', 'credit']
    cl_display_cols = [c for c in cl_display_cols if c in cl_ann.columns]
    cl_display_cols += ['_remark', '_match_ref']
    ncols_cl = len(cl_display_cols)

    # Row 1: Title
    ws_cl.merge_cells(f'A1:{get_column_letter(ncols_cl)}1')
    title_cl = ws_cl['A1']
    title_cl.value = f'📗 {CL} — CUSTOMER LEDGER WITH REMARKS'
    title_cl.font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
    title_cl.fill = mk_fill('1A6B45')
    title_cl.alignment = Alignment(horizontal='center', vertical='center')
    ws_cl.row_dimensions[1].height = 28

    # Row 2: SUBTOTAL at top (formulas filled after data)
    for c_idx in range(1, ncols_cl + 1):
        cell = ws_cl.cell(row=2, column=c_idx)
        cell.fill = mk_fill('1A4A35')
        cell.border = border
        cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_cl.cell(row=2, column=1).value = 'SUBTOTAL ▲'
    ws_cl.row_dimensions[2].height = 22

    # Row 3: Headers
    cl_hmap = {'doc_date':'Doc Date','doc_no':'Doc No','doc_type':'Doc Type',
               'debit':'Debit (LC)','credit':'Credit (LC)',
               '_remark':'Remark','_match_ref':'Matched With'}
    headers = [cl_hmap.get(c, c) for c in cl_display_cols]
    style_header(ws_cl, headers, row=3, color='1A6B45')

    debit_col_cl  = (cl_display_cols.index('debit')  + 1) if 'debit'  in cl_display_cols else None
    credit_col_cl = (cl_display_cols.index('credit') + 1) if 'credit' in cl_display_cols else None

    # Data starts at row 4
    CL_DATA_START = 4
    for r_idx, (_, row) in enumerate(cl_ann[cl_display_cols].iterrows(), CL_DATA_START):
        remark = str(row.get('_remark', ''))
        fill = COLORS['unmatched_fill'] if 'Unmatched' in remark else COLORS['matched_fill']
        for c_idx, col in enumerate(cl_display_cols, 1):
            val = row[col]
            cell = write_cell(ws_cl, r_idx, c_idx, val)
            cell.fill = mk_fill(fill)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            cell.font = Font(name='Calibri', size=9)
        ws_cl.row_dimensions[r_idx].height = 18

    cl_data_end = CL_DATA_START + len(cl_ann) - 1

    # Fill SUBTOTAL formulas at row 2
    for col_i in [debit_col_cl, credit_col_cl]:
        if col_i:
            cl_l = get_column_letter(col_i)
            cell = ws_cl.cell(row=2, column=col_i,
                              value=f'=SUBTOTAL(9,{cl_l}{CL_DATA_START}:{cl_l}{cl_data_end})')
            cell.number_format = '#,##0.00'
            cell.fill = mk_fill('1A4A35')
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')
    auto_width(ws_cl)

    # ══════════════════════════════════════════
    # DETAIL SHEETS WITH SUBTOTALS
    # ══════════════════════════════════════════
    def write_sheet(title, data, hdr_color, is_vl_sheet=None):
        """Write an annexure sheet with: subtotal at top, color per column (VL=blue, CL=teal), names instead of VL/CL."""
        if not data:
            return
        ws = wb.create_sheet(title[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A3'  # row 1=title, row 2=subtotal, row3=header, data from row4

        df = pd.DataFrame(data)
        # Replace VL/CL labels in column names with actual names
        rename_map = {}
        for col in df.columns:
            new_col = col.replace(' VL ', f' {VL} ').replace(' CL ', f' {CL} ')
            new_col = new_col.replace('VL ', f'{VL} ').replace('CL ', f'{CL} ')
            new_col = new_col.replace(' VL', f' {VL}').replace(' CL', f' {CL}')
            if new_col != col:
                rename_map[col] = new_col
        if rename_map:
            df = df.rename(columns=rename_map)
        hdrs = list(df.columns)
        ncols = len(hdrs)

        # Row 1: Title band
        ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
        t = ws['A1']
        t.value = title.replace('VL', VL).replace('CL', CL)
        t.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        t.fill = mk_fill(hdr_color)
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        # Row 2: SUBTOTAL (filled after data)
        SUBT_ROW = 2
        for c_idx in range(1, ncols+1):
            cell = ws.cell(row=SUBT_ROW, column=c_idx)
            cell.fill = mk_fill('2A3A5A')
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=SUBT_ROW, column=1).value = 'SUBTOTAL ▲'
        ws.row_dimensions[SUBT_ROW].height = 22

        # Row 3: Column headers with VL=blue, CL=teal color per column
        for c_idx, h in enumerate(hdrs, 1):
            hu = h.upper()
            if VL.upper()[:4] in hu or 'VL' in hu:
                hc = VL_COLOR
            elif CL.upper()[:4] in hu or 'CL' in hu:
                hc = CL_COLOR
            else:
                hc = hdr_color
            cell = ws.cell(row=3, column=c_idx, value=h)
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=9)
            cell.fill = mk_fill(hc)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[3].height = 30

        # Data rows from row 4
        DATA_START_WS = 4
        for r, (_, row) in enumerate(df.iterrows(), DATA_START_WS):
            for c_idx, (h, val) in enumerate(zip(hdrs, row.values), 1):
                hu = h.upper()
                # Per-column color fill
                if VL.upper()[:4] in hu or 'VL' in hu:
                    fill = VL_LIGHT if r % 2 == 0 else 'F0F4FF'
                elif CL.upper()[:4] in hu or 'CL' in hu:
                    fill = CL_LIGHT if r % 2 == 0 else 'F0FAF5'
                else:
                    fill = 'F8F9FB' if r % 2 == 0 else 'FFFFFF'
                cell = write_cell(ws, r, c_idx, val)
                cell.fill = mk_fill(fill)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=(c_idx == 1))
                cell.font = Font(name='Calibri', size=9)
            ws.row_dimensions[r].height = 18

        data_end_ws = DATA_START_WS + len(data) - 1

        # Fill subtotal formulas at row 2
        for c_idx, h in enumerate(hdrs, 1):
            if any(k in h.lower() for k in ['debit','credit','amount','value','reversal','original']):
                cl_l = get_column_letter(c_idx)
                cell = ws.cell(row=SUBT_ROW, column=c_idx,
                               value=f'=SUBTOTAL(9,{cl_l}{DATA_START_WS}:{cl_l}{data_end_ws})')
                cell.number_format = '#,##0.00'
                cell.fill = mk_fill('2A3A5A')
                cell.font = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
                cell.border = border
                cell.alignment = Alignment(horizontal='right', vertical='center')

        auto_width(ws)

    vl6 = VL[:10]; cl6 = CL[:10]
    write_sheet(f'Inv-Matched',                   results['invoice_matched'],         '1A6B45')
    write_sheet(f'Inv-Unmatch-{vl6}',             results['invoice_unmatched_vl'],    'A32035', is_vl_sheet=True)
    write_sheet(f'Inv-Unmatch-{cl6}',             results['invoice_unmatched_cl'],    'A32035', is_vl_sheet=False)
    write_sheet(f'CN-Unmatch-{vl6}',              results['cn_unmatched_vl'],         'B85C00', is_vl_sheet=True)
    write_sheet(f'DN-CN-Matched',                 results['dn_matched'],              '1A6B45')
    write_sheet(f'DN-Unmatch-{vl6}',              results['dn_unmatched_vl'],         'A32035', is_vl_sheet=True)
    write_sheet(f'DN-Unmatch-{cl6}',              results['dn_unmatched_cl'],         'A32035', is_vl_sheet=False)
    write_sheet(f'Coll-Matched',                  results['collection_matched'],      '1A6B45')
    write_sheet(f'Coll-Unmatch-{vl6}',            results['collection_unmatched_vl'], 'A32035', is_vl_sheet=True)
    write_sheet(f'Coll-Unmatch-{cl6}',            results['collection_unmatched_cl'], 'A32035', is_vl_sheet=False)
    write_sheet(f'AnnexA-CrossLedger',            results['reversal_cross_ledger'],   'B85C00')
    write_sheet(f'AnnexB-{vl6}-Internal',         results['reversal_vl_internal'],    '7B5EA7')
    rev_mis  = [r for r in results['reversal_unmatched'] if r.get('Reason','') == 'Amount Mismatch']
    rev_miss = [r for r in results['reversal_unmatched'] if r.get('Reason','') != 'Amount Mismatch']
    write_sheet(f'AnnexC1-AmtMismatch',           rev_mis,                            'A32035')
    write_sheet(f'AnnexC2-MissingOrig',           rev_miss,                           'A32035')

    # ══════════════════════════════════════════
    # RECON STATEMENT SHEET
    # ══════════════════════════════════════════
    ws_rs = wb.create_sheet('Recon Statement', 1)
    ws_rs.sheet_view.showGridLines = False
    ws_rs.column_dimensions['A'].width = 62
    ws_rs.column_dimensions['B'].width = 22
    ws_rs.column_dimensions['C'].width = 28

    def ssum(lst, key):
        try: return sum(float(d.get(key, 0) or 0) for d in lst)
        except: return 0.0

    inv_un_vl_v  = ssum(results['invoice_unmatched_vl'],   'Debit') + ssum(results['invoice_unmatched_vl'],   'Credit')
    inv_un_cl_v  = ssum(results['invoice_unmatched_cl'],   'Debit') + ssum(results['invoice_unmatched_cl'],   'Credit')
    dn_un_cl_v   = ssum(results['dn_unmatched_cl'],        'Debit') + ssum(results['dn_unmatched_cl'],        'Credit')
    col_un_vl_v  = ssum(results['collection_unmatched_vl'],'Debit') + ssum(results['collection_unmatched_vl'],'Credit')
    col_un_cl_v  = ssum(results['collection_unmatched_cl'],'Debit') + ssum(results['collection_unmatched_cl'],'Credit')
    # Credit notes unmatched in VL (Saleable Return, Non-Saleable, Credit Note)
    cn_un_vl_v   = ssum(results.get('cn_unmatched_vl', []), 'Debit') + ssum(results.get('cn_unmatched_vl', []), 'Credit')
    vl_close     = results.get('vl_closing') or 0.0
    cl_close     = results.get('cl_closing') or 0.0

    # Sheet name references (truncated to match actual sheet tab names)
    vl6 = VL[:10]; cl6 = CL[:10]
    src = {
        'inv_vl':  f'Inv-Unmatch-{vl6}'[:31],
        'inv_cl':  f'Inv-Unmatch-{cl6}'[:31],
        'cn_vl':   f'CN-Unmatch-{vl6}'[:31],
        'dn_cl':   f'DN-Unmatch-{cl6}'[:31],
        'col_vl':  f'Coll-Unmatch-{vl6}'[:31],
        'col_cl':  f'Coll-Unmatch-{cl6}'[:31],
        'vl_ledger': f'{VL[:18]} Vendor Ledger'[:31],
        'cl_ledger': f'{CL[:18]} Customer Ledger'[:31],
    }

    # Each row: (label, amount, row_type, source_sheet_name)
    rs_rows = [
        ('Particular',                                              'Amount (₹)',    'H',  'Source Sheet'),
        (f'Balance as per {VL} Books (A)',                          vl_close,        'VL', src['vl_ledger']),
        (f'Less:  Tax Invoice in {VL} not available in {CL}',      -inv_un_vl_v,    'L',  src['inv_vl']),
        (f'Add:   Credit Note in {VL} not available in {CL}',       cn_un_vl_v,     'A',  src['cn_vl']),
        (f'Less:  Debit Notes in {CL} not available in {VL}',      -dn_un_cl_v,     'L',  src['dn_cl']),
        (f'Add:   Tax Invoice in {CL} not available in {VL}',       inv_un_cl_v,    'A',  src['inv_cl']),
        (f'Add:   Payment not available in {CL}',                   col_un_vl_v,    'A',  src['col_vl']),
        (f'Less:  Payment not available in {VL}',                  -col_un_cl_v,    'L',  src['col_cl']),
        ('',                                                         '',             'B',  ''),
        (f'Net Balance as per {VL} Books — B',                     '=SUM(B2:B8)',   'T',  ''),
        ('',                                                         '',             'B',  ''),
        (f'Balance as per {CL} Books — C',                          cl_close,       'CL', src['cl_ledger']),
        ('',                                                         '',             'B',  ''),
        ('Unreconciled Difference (B - C)',                         '=B10-B12',     'D',  ''),
        ('(This value should be zero after all the adjustments)',    '',             'N',  ''),
    ]

    fill_map = {
        'H':  ('1C2130', 'E8ECF4'),
        'VL': ('1A3A6B', 'FFFFFF'),
        'L':  ('FFE8E8', 'A32035'),
        'A':  ('E8F5EE', '1A6B45'),
        'B':  ('FFFFFF', 'FFFFFF'),
        'T':  ('1A3A6B', 'FFFFFF'),
        'CL': ('1A6B45', 'FFFFFF'),
        'D':  ('CC0000', 'FFFFFF'),
        'N':  ('CC0000', 'FFFFFF'),
    }

    for r_ptr, (label, amount, rtype, src_sheet) in enumerate(rs_rows, 1):
        bg, fg = fill_map.get(rtype, ('FFFFFF', '000000'))
        is_bold = rtype in ('H', 'VL', 'T', 'CL', 'D')

        # Column A: Particular
        cell_a = ws_rs.cell(row=r_ptr, column=1, value=label)
        cell_a.fill = PatternFill(fill_type='solid', fgColor=bg)
        cell_a.font = Font(name='Calibri', size=10, color=fg, bold=is_bold)
        cell_a.border = border
        cell_a.alignment = Alignment(vertical='center')

        # Column B: Amount
        cell_b = ws_rs.cell(row=r_ptr, column=2, value=amount)
        cell_b.fill = PatternFill(fill_type='solid', fgColor=bg)
        cell_b.font = Font(name='Calibri', size=10, color=fg, bold=is_bold)
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal='right', vertical='center')
        if isinstance(amount, (int, float)):
            cell_b.number_format = '#,##0.00'

        # Column C: Source Sheet hyperlink
        cell_c = ws_rs.cell(row=r_ptr, column=3, value=f'→ {src_sheet}' if src_sheet else '')
        cell_c.fill = PatternFill(fill_type='solid', fgColor=bg)
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal='center', vertical='center')
        if src_sheet and rtype not in ('H', 'B', 'T', 'D', 'N'):
            cell_c.font = Font(name='Calibri', size=9, color='1A6BCC',
                               bold=True, underline='single')
        else:
            cell_c.font = Font(name='Calibri', size=9, color=fg)

        ws_rs.row_dimensions[r_ptr].height = 22 if rtype != 'B' else 8

    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# UI HELPERS
# ─────────────────────────────────────────────

def display_df(df):
    if df is None or (isinstance(df, pd.DataFrame) and df.empty) or (isinstance(df, list) and len(df) == 0):
        st.info("No records in this category.")
        return
    if isinstance(df, list):
        df = pd.DataFrame(df)
    df = df.copy()
    for col in df.columns:
        if 'date' in col.lower() or 'Date' in col:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d-%b-%Y').fillna('')
    st.dataframe(df, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────

def fmt_inr(val):
    """Format number as Indian currency."""
    try:
        v = float(val)
        return f"₹{v:,.2f}"
    except:
        return "₹0.00"

def safe_sum(lst, key):
    """Sum a key across a list of dicts."""
    try:
        return sum(float(d.get(key, 0) or 0) for d in lst)
    except:
        return 0.0

# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────

def main():
    # ── Step 0: Vendor / Customer Name Entry ──
    # Collect names before anything else is shown
    if 'vname' not in st.session_state:
        st.session_state['vname'] = ''
    if 'cname' not in st.session_state:
        st.session_state['cname'] = ''

    # Header
    st.markdown("""
    <div class="recon-header">
        <div>
            <div class="recon-logo">⚖️ Ledger Reconciliation</div>
            <div class="recon-subtitle">Vendor · Customer Ledger Reconciliation · For Indian CAs &amp; CFOs</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Name inputs always visible at top
    nc1, nc2, nc3 = st.columns([2, 2, 1])
    with nc1:
        vname = st.text_input("🏭 Vendor Name", value=st.session_state['vname'],
                               placeholder="e.g. ABC Suppliers Pvt. Ltd.")
        if vname:
            st.session_state['vname'] = vname
    with nc2:
        cname = st.text_input("🏢 Customer Name", value=st.session_state['cname'],
                               placeholder="e.g. XYZ Traders Ltd.")
        if cname:
            st.session_state['cname'] = cname

    vname = st.session_state.get('vname', 'Vendor') or 'Vendor'
    cname = st.session_state.get('cname', 'Customer') or 'Customer'

    VL = vname   # use actual names everywhere
    CL = cname

    # Define tolerance early so it's available in instructions before sidebar renders
    tolerance = st.session_state.get('tolerance_val', 1.0)

    # ── How to Use Instructions ──
    with st.expander("📖 How to Reconcile — Full Instructions", expanded=False):
        st.markdown(f"""
        <div class="info-box" style="line-height:1.9;">
        <b style="font-size:1rem;color:#4f8eff;">Step-by-Step Reconciliation Guide</b><br><br>

        <b>1. Enter Names</b><br>
        Type the Vendor name (your supplier's name as appearing in their books) and Customer name (your company name as in your books). These names will appear on all reports and Excel sheets.<br><br>

        <b>2. Prepare Your Ledger Files</b><br>
        Export both ledgers as <b>.xlsx</b> from your ERP (Tally, SAP, Oracle, Zoho, QuickBooks, etc.).<br>
        &nbsp;&nbsp;• <b>Vendor Ledger (VL):</b> The ledger your vendor has shared — shows how they have recorded transactions with you. Must have: Date, Doc No, Doc Type, Debit, Credit, Closing Balance.<br>
        &nbsp;&nbsp;• <b>Customer Ledger (CL):</b> Your own ledger — shows the same party's account in your books. Must have: Date, Doc No, Doc Type, Debit, Credit, Closing Balance.<br>
        &nbsp;&nbsp;• The app auto-detects column headers. If columns are not detected, use the <b>Column Mapping</b> section below to assign them manually.<br>
        &nbsp;&nbsp;• The <b>Closing Balance</b> is read from the last row of the Closing/Balance column automatically.<br><br>

        <b>3. Understand What Gets Matched</b><br>
        &nbsp;&nbsp;• <b>Tax Invoices / Sales Invoices</b> → Matched by exact Document Number between VL and CL.<br>
        &nbsp;&nbsp;• <b>Credit Notes, Saleable Returns, Non-Saleable Returns</b> (in VL) → Matched against Discount Debit Notes and PRN entries in CL.<br>
        &nbsp;&nbsp;• <b>Debit Notes</b> → Matched by Document Number, then by Period + Amount (within ₹{tolerance} tolerance).<br>
        &nbsp;&nbsp;• <b>Collections / Payments</b> → Matched by UTR number extracted from Particulars/Doc No, then by Period + Amount.<br>
        &nbsp;&nbsp;• <b>Complete Reversals</b> in VL → The original invoice is traced using the Particulars column. Three outcomes: (A) original also in CL (needs review), (B) purely internal VL reversal, (C) original not found or amount mismatch.<br><br>

        <b>4. What the Reconciliation Statement Shows</b><br>
        The statement starts from the <b>VL Closing Balance</b>, then adjusts for all differences to arrive at a Net Balance (B). This is then compared against the <b>CL Closing Balance</b> (C). The difference (B–C) should be zero once all items are investigated and resolved.<br>
        &nbsp;&nbsp;• <b>Less: Invoices in VL not in CL</b> — Invoices your vendor shows but you haven't recorded yet.<br>
        &nbsp;&nbsp;• <b>Add: Credit Notes in VL not in CL</b> — Returns/credits your vendor shows but not yet in your books.<br>
        &nbsp;&nbsp;• <b>Less: Debit Notes in CL not in VL</b> — Deductions you've recorded (discounts/claims) not yet acknowledged by vendor.<br>
        &nbsp;&nbsp;• <b>Add: Invoices in CL not in VL</b> — Invoices in your books the vendor hasn't reflected yet.<br>
        &nbsp;&nbsp;• <b>Add/Less: Payments</b> — Payments recorded in one ledger but not yet the other.<br><br>

        <b>5. Download the Report</b><br>
        Click the green <b>⬇️ Download Report</b> button to get the full Excel workbook with 18 sheets — Summary, Recon Statement, both ledgers with colour-coded remarks, and 14 annexure sheets covering every matched/unmatched category in detail.<br><br>

        <b>6. Action Checklist After Reconciliation</b><br>
        &nbsp;&nbsp;✅ Matched items → No action needed.<br>
        &nbsp;&nbsp;🔴 Unmatched Invoices → Check if invoice was raised/received; book it if missing.<br>
        &nbsp;&nbsp;🟡 Credit Notes in VL not in CL → Raise a corresponding debit note or book the credit note in your books.<br>
        &nbsp;&nbsp;⚠️ Reversal Annexure A → Invoice reversed by vendor but still open in your books — investigate and reverse or close.<br>
        &nbsp;&nbsp;❌ Reversal Annexure C → Amount mismatch or missing original — get vendor clarification.<br>
        &nbsp;&nbsp;💰 Unmatched Collections → Confirm UTR with bank; match manually if needed.
        </div>
        """, unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown(f"### ⚙️ Configuration")
        st.markdown("---")
        tolerance = st.number_input(
            "Amount Tolerance (₹)",
            min_value=0.0, max_value=100.0, value=1.0, step=0.5,
            help="Max difference allowed when matching by amount"
        )
        st.session_state['tolerance_val'] = tolerance
        st.markdown("---")
        st.markdown("### 📋 5-Step Matching Logic")
        st.markdown(f"""
        <div class="info-box">
        <b>Step 1 — Complete Reversals</b><br>
        Finds VL rows with type "Complete Reversal". Looks up the original invoice from Particulars. Three sub-cases:<br>
        &nbsp;&nbsp;A: Reversed in VL <i>and</i> present in CL → ⚠️ Needs Review<br>
        &nbsp;&nbsp;B: Reversed in VL, NOT in CL → ✅ Internal only<br>
        &nbsp;&nbsp;C: No original found / amount mismatch → ❌ Unmatched<br><br>
        <b>Step 2 — Invoice Matching</b><br>
        Matches VL Tax Invoices vs CL Invoices by exact Document Number. Excludes Credit Notes, DN, Collections.<br><br>
        <b>Step 2B — Credit Notes vs Discount DN/PRN</b><br>
        VL Credit Notes (incl. Saleable Return, Non-Saleable) matched against CL Discount Debit Notes and PRN entries. First by Doc No, then by Period+Amount.<br><br>
        <b>Step 3 — Debit Note Matching</b><br>
        Matches by Doc Number first, then Period+Amount (within ₹{tolerance} tolerance).<br><br>
        <b>Step 4 — Collection Matching</b><br>
        Extracts UTR from Particulars/Doc No. Matches by UTR first, then Period+Amount.<br><br>
        <b>Step 5 — Unmatched</b><br>
        Remaining rows go to separate unmatched buckets for Invoices, Credit Notes, Debit Notes, and Collections.
        </div>
        """, unsafe_allow_html=True)
        st.markdown("---")
        st.markdown(f"""
        <div class="info-box" style="font-size:0.72rem;">
        <b>Color Code:</b><br>
        <span style="color:#4f8eff">■ {VL}</span> (Blue)<br>
        <span style="color:#00d4aa">■ {CL}</span> (Teal)<br>
        <span style="color:#00d4aa">■ Matched</span> · <span style="color:#ff4d6d">■ Unmatched</span> · <span style="color:#ff8c42">■ Reversal</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Column Mapping Helper ──
    with st.expander("⚙️ Column Mapping  |  📥 Download Sample Format", expanded=False):
        st.markdown("""
        <div class="info-box">
        <b>How it works:</b> Upload any Excel ledger file — the app auto-detects your columns by keywords.
        If your columns are not detected correctly, use the mapping dropdowns below to manually assign them.
        You can also download a sample format to understand the expected structure.
        </div>
        """, unsafe_allow_html=True)

        samp1, samp2 = st.columns(2)

        vl_sample_buf = BytesIO()
        pd.DataFrame({
            'Doc Date':       ['01-Apr-2025', '05-Apr-2025', '10-Apr-2025'],
            'Doc No':         ['INV/001',     'INV/002',     'PAY/001'],
            'Doc Type Name':  ['Tax Invoice', 'Tax Invoice', 'Payment'],
            'Particulars':    ['Sale of goods','Sale of goods','NEFT payment received'],
            'Debit':          [10000, 15000, 0],
            'Credit':         [0, 0, 8000],
            'Closing Balance':[10000, 25000, 17000],
        }).to_excel(vl_sample_buf, index=False)
        samp1.download_button('⬇️ Sample Vendor Ledger',
            data=vl_sample_buf.getvalue(),
            file_name='Sample_Vendor_Ledger.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True, key='sample_vl')

        cl_sample_buf = BytesIO()
        pd.DataFrame({
            'Document Date':  ['01-Apr-2025', '05-Apr-2025', '10-Apr-2025'],
            'Document Type':  ['Tax Invoice', 'Tax Invoice', 'Payment'],
            'Document No':    ['INV/001',     'INV/002',     'PAY/001'],
            'Debit (LC)':     [0, 0, 8000],
            'Credit (LC)':    [10000, 15000, 0],
            'Closing Balance':[-10000, -25000, -17000],
        }).to_excel(cl_sample_buf, index=False)
        samp2.download_button('⬇️ Sample Customer Ledger',
            data=cl_sample_buf.getvalue(),
            file_name='Sample_Customer_Ledger.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True, key='sample_cl')

        st.markdown("---")
        st.markdown("**Manual Column Mapping** — only needed if auto-detection is incorrect")

        # Show mapping UI only if files are already uploaded and parsed
        if 'vl_parsed' in st.session_state and 'cl_parsed' in st.session_state:
            vl_cols_all = [c for c in st.session_state['vl_parsed'].columns if not c.startswith('_')]
            cl_cols_all = [c for c in st.session_state['cl_parsed'].columns if not c.startswith('_')]
            NONE = '(auto)'

            m1, m2 = st.columns(2)
            with m1:
                st.markdown(f'<span class="section-tag tag-vl">📘 {VL} — Vendor Ledger Columns</span>', unsafe_allow_html=True)
                vl_map_date = st.selectbox('Doc Date column', [NONE]+vl_cols_all, key='vl_map_date')
                vl_map_no   = st.selectbox('Doc No column',   [NONE]+vl_cols_all, key='vl_map_no')
                vl_map_type = st.selectbox('Doc Type column', [NONE]+vl_cols_all, key='vl_map_type')
                vl_map_part = st.selectbox('Particulars column', [NONE]+vl_cols_all, key='vl_map_part')
                vl_map_deb  = st.selectbox('Debit column',    [NONE]+vl_cols_all, key='vl_map_deb')
                vl_map_cred = st.selectbox('Credit column',   [NONE]+vl_cols_all, key='vl_map_cred')
                vl_map_clos = st.selectbox('Closing Balance column', [NONE]+vl_cols_all, key='vl_map_clos')
            with m2:
                st.markdown(f'<span class="section-tag tag-cl">📗 {CL} — Customer Ledger Columns</span>', unsafe_allow_html=True)
                cl_map_date = st.selectbox('Doc Date column', [NONE]+cl_cols_all, key='cl_map_date')
                cl_map_no   = st.selectbox('Doc No column',   [NONE]+cl_cols_all, key='cl_map_no')
                cl_map_type = st.selectbox('Doc Type column', [NONE]+cl_cols_all, key='cl_map_type')
                cl_map_deb  = st.selectbox('Debit column',    [NONE]+cl_cols_all, key='cl_map_deb')
                cl_map_cred = st.selectbox('Credit column',   [NONE]+cl_cols_all, key='cl_map_cred')
                cl_map_clos = st.selectbox('Closing Balance column', [NONE]+cl_cols_all, key='cl_map_clos')

            if st.button('✅ Apply Manual Column Mapping', key='apply_col_map'):
                def apply_map(df, mapping):
                    """Re-map columns based on user selections."""
                    df = df.copy()
                    for target, src in mapping.items():
                        if src and src != NONE and src in df.columns and src != target:
                            df[target] = df[src]
                    # Re-run numeric conversions
                    for nc in ['debit','credit','closing']:
                        if nc in df.columns:
                            df[nc] = pd.to_numeric(df[nc], errors='coerce').fillna(0 if nc != 'closing' else np.nan)
                    if 'doc_no' in df.columns:
                        df['doc_no_clean'] = df['doc_no'].apply(clean_doc_number)
                    if 'doc_date' in df.columns:
                        df['doc_date'] = pd.to_datetime(df['doc_date'], errors='coerce', dayfirst=True)
                        df['period'] = df['doc_date'].apply(get_period)
                    return df

                vl_remap = {'doc_date':vl_map_date,'doc_no':vl_map_no,'doc_type':vl_map_type,
                            'particulars':vl_map_part,'debit':vl_map_deb,'credit':vl_map_cred,'closing':vl_map_clos}
                cl_remap = {'doc_date':cl_map_date,'doc_no':cl_map_no,'doc_type':cl_map_type,
                            'debit':cl_map_deb,'credit':cl_map_cred,'closing':cl_map_clos}

                st.session_state['vl_parsed'] = apply_map(st.session_state['vl_parsed'], vl_remap)
                st.session_state['cl_parsed'] = apply_map(st.session_state['cl_parsed'], cl_remap)
                st.session_state.pop('results', None)
                st.session_state.pop('excel_data', None)
                st.success('✅ Column mapping applied. Click Run Reconciliation.')
        else:
            st.info('Upload both ledger files first to use manual column mapping.')

    # File upload with vendor/customer color coding
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f'<span class="section-tag tag-vl">📘 {VL.upper()} — VENDOR LEDGER</span>', unsafe_allow_html=True)
        vl_file = st.file_uploader(f"Upload {VL} Ledger (.xlsx / .xls)", type=['xlsx', 'xls'], key='vl')
    with col2:
        st.markdown(f'<span class="section-tag tag-cl">📗 {CL.upper()} — CUSTOMER LEDGER</span>', unsafe_allow_html=True)
        cl_file = st.file_uploader(f"Upload {CL} Ledger (.xlsx / .xls)", type=['xlsx', 'xls'], key='cl')

    if not vl_file or not cl_file:
        st.markdown(f"""
        <div class="info-box" style="margin-top:2rem; border-left-color: #4f8eff;">
        📂 Upload both ledger files above to begin reconciliation.<br>
        Enter <b>{VL}</b> and <b>{CL}</b> names above for personalised output.
        </div>
        """, unsafe_allow_html=True)
        # Clear stale results if files removed
        st.session_state.pop('results', None)
        st.session_state.pop('excel_data', None)
        st.session_state.pop('vl_parsed', None)
        st.session_state.pop('cl_parsed', None)
        return

    # ── Cache parsed files — only re-parse when file changes ──
    vl_file_id = (vl_file.name, vl_file.size)
    cl_file_id = (cl_file.name, cl_file.size)

    if st.session_state.get('vl_file_id') != vl_file_id or 'vl_parsed' not in st.session_state:
        with st.spinner(f"Parsing {VL} ledger..."):
            try:
                vl_df, vl_close = load_vendor_ledger(vl_file)
                st.session_state['vl_parsed'] = vl_df
                st.session_state['vl_closing_val'] = vl_close
                st.session_state['vl_file_id'] = vl_file_id
                st.session_state.pop('results', None)
                st.session_state.pop('excel_data', None)
            except Exception as e:
                st.error(f"Error reading {VL} file: {e}")
                return

    if st.session_state.get('cl_file_id') != cl_file_id or 'cl_parsed' not in st.session_state:
        with st.spinner(f"Parsing {CL} ledger..."):
            try:
                cl_df, cl_close = load_customer_ledger(cl_file)
                st.session_state['cl_parsed'] = cl_df
                st.session_state['cl_closing_val'] = cl_close
                st.session_state['cl_file_id'] = cl_file_id
                st.session_state.pop('results', None)
                st.session_state.pop('excel_data', None)
            except Exception as e:
                st.error(f"Error reading {CL} file: {e}")
                return

    vl = st.session_state['vl_parsed']
    cl = st.session_state['cl_parsed']

    # Closing balances — stored separately in session_state (DataFrame attributes don't persist)
    vl_closing = st.session_state.get('vl_closing_val')
    if vl_closing is None:
        # Fallback: try from the filtered df closing column (last non-null value)
        vl_closing = vl['closing'].dropna().iloc[-1] if 'closing' in vl.columns and not vl['closing'].dropna().empty else None

    cl_closing = st.session_state.get('cl_closing_val')
    if cl_closing is None:
        cl_closing = cl['closing'].dropna().iloc[-1] if 'closing' in cl.columns and not cl['closing'].dropna().empty else None

    st.success(f"✅ {VL}: **{len(vl)}** rows  ·  {CL}: **{len(cl)}** rows")

    # Show auto-detected column mapping so user can verify
    vl_detected = {k: k for k in ['doc_date','doc_no','doc_type','debit','credit','closing'] if k in vl.columns}
    cl_detected = {k: k for k in ['doc_date','doc_no','doc_type','debit','credit','closing'] if k in cl.columns}
    missing_vl = [k for k in ['doc_date','doc_no','doc_type','debit','credit'] if k not in vl.columns]
    missing_cl = [k for k in ['doc_date','doc_no','doc_type','debit','credit'] if k not in cl.columns]
    if missing_vl or missing_cl:
        st.warning(f"⚠️ Some columns could not be auto-detected. "
                   f"{'VL missing: ' + ', '.join(missing_vl) if missing_vl else ''} "
                   f"{'CL missing: ' + ', '.join(missing_cl) if missing_cl else ''}. "
                   f"Use **Column Mapping** above to fix this.")

    # Show both closing balances side by side
    cb1, cb2 = st.columns(2)
    with cb1:
        if vl_closing is not None:
            st.info(f"📘 **{VL} Closing Balance:** {fmt_inr(vl_closing)}")
        else:
            st.info(f"📘 **{VL} Closing Balance:** Not detected in file")
    with cb2:
        if cl_closing is not None:
            st.info(f"📗 **{CL} Closing Balance:** {fmt_inr(cl_closing)}")
        else:
            st.info(f"📗 **{CL} Closing Balance:** Not detected in file")

    with st.expander("👁 Preview Parsed Data"):
        pc1, pc2 = st.columns(2)
        with pc1:
            st.markdown(f'<span class="section-tag tag-vl">{VL} — VENDOR LEDGER</span>', unsafe_allow_html=True)
            display_cols = [c for c in ['doc_date', 'doc_no', 'doc_type', 'particulars', 'debit', 'credit', 'closing'] if c in vl.columns]
            st.dataframe(vl[display_cols].head(10), use_container_width=True, hide_index=True)
        with pc2:
            st.markdown(f'<span class="section-tag tag-cl">{CL} — CUSTOMER LEDGER</span>', unsafe_allow_html=True)
            display_cols = [c for c in ['doc_date', 'doc_no', 'doc_type', 'debit', 'credit'] if c in cl.columns]
            st.dataframe(cl[display_cols].head(10), use_container_width=True, hide_index=True)

    if st.button("▶ Run Reconciliation", use_container_width=False):
        with st.spinner("Running reconciliation engine..."):
            results = run_reconciliation(vl, cl, tolerance=tolerance)

        def safe_records(df):
            """Convert DataFrame to list of dicts safely — handles any column types."""
            if not isinstance(df, pd.DataFrame):
                return df  # already a list or other type
            d = df.copy()
            for col in list(d.columns):
                try:
                    if col.startswith('_'):
                        d[col] = d[col].astype(str).replace('nan', '').replace('None', '')
                    elif hasattr(d[col], 'dtype') and d[col].dtype == object:
                        d[col] = d[col].fillna('').astype(str)
                    elif hasattr(d[col], 'dtype') and str(d[col].dtype).startswith('datetime'):
                        d[col] = d[col].astype(str)
                except Exception:
                    try:
                        d[col] = d[col].fillna('').astype(str)
                    except Exception:
                        pass
            return d.to_dict('records')

        results['vl_annotated'] = safe_records(results['vl_annotated'])
        results['cl_annotated'] = safe_records(results['cl_annotated'])
        results['vl_closing'] = float(vl_closing) if vl_closing is not None else None
        results['cl_closing'] = float(cl_closing) if cl_closing is not None else None
        results['vl_name'] = VL
        results['cl_name'] = CL
        st.session_state['results'] = results
        # Clear cached excel so it regenerates fresh with new results
        st.session_state.pop('excel_data', None)
        st.session_state.pop('excel_key', None)

    if 'results' not in st.session_state:
        return

    results = st.session_state['results']
    VL = results.get('vl_name', vname) or vname
    CL = results.get('cl_name', cname) or cname
    vl_ann_df = pd.DataFrame(results['vl_annotated'])
    cl_ann_df  = pd.DataFrame(results['cl_annotated'])
    vl_closing_val = results.get('vl_closing')
    cl_closing_val = results.get('cl_closing')

    # ── Compute values for summary ──
    cn_matched      = [r for r in results['dn_matched'] if 'Credit Note' in str(r.get('Match Type', ''))]
    dn_only_matched = [r for r in results['dn_matched'] if 'Credit Note' not in str(r.get('Match Type', ''))]

    inv_matched_cnt  = len(results['invoice_matched'])
    inv_un_vl_cnt    = len(results['invoice_unmatched_vl'])
    inv_un_cl_cnt    = len(results['invoice_unmatched_cl'])
    dn_matched_cnt   = len(dn_only_matched)
    dn_un_vl_cnt     = len(results['dn_unmatched_vl'])
    dn_un_cl_cnt     = len(results['dn_unmatched_cl'])
    cn_matched_cnt   = len(cn_matched)
    col_matched_cnt  = len(results['collection_matched'])
    col_un_vl_cnt    = len(results['collection_unmatched_vl'])
    col_un_cl_cnt    = len(results['collection_unmatched_cl'])
    rev_cross_cnt    = len(results['reversal_cross_ledger'])
    rev_int_cnt      = len(results['reversal_vl_internal'])
    rev_un_cnt       = len(results['reversal_unmatched'])
    rev_amt_mis_cnt  = len([r for r in results['reversal_unmatched'] if r.get('Reason','') == 'Amount Mismatch'])
    rev_miss_cnt     = rev_un_cnt - rev_amt_mis_cnt

    # Values (credit/debit sums)
    inv_matched_val  = safe_sum(results['invoice_matched'], 'VL Debit') + safe_sum(results['invoice_matched'], 'VL Credit')
    inv_un_vl_val    = safe_sum(results['invoice_unmatched_vl'], 'Debit') + safe_sum(results['invoice_unmatched_vl'], 'Credit')
    inv_un_cl_val    = safe_sum(results['invoice_unmatched_cl'], 'Debit') + safe_sum(results['invoice_unmatched_cl'], 'Credit')
    dn_matched_val   = safe_sum(dn_only_matched, 'VL Debit') + safe_sum(dn_only_matched, 'VL Credit')
    dn_un_vl_val     = safe_sum(results['dn_unmatched_vl'], 'Debit') + safe_sum(results['dn_unmatched_vl'], 'Credit')
    dn_un_cl_val     = safe_sum(results['dn_unmatched_cl'], 'Debit') + safe_sum(results['dn_unmatched_cl'], 'Credit')
    cn_matched_val   = safe_sum(cn_matched, 'VL Debit') + safe_sum(cn_matched, 'VL Credit')
    col_matched_val  = safe_sum(results['collection_matched'], 'VL Amount')
    col_un_vl_val    = safe_sum(results['collection_unmatched_vl'], 'Debit') + safe_sum(results['collection_unmatched_vl'], 'Credit')
    col_un_cl_val    = safe_sum(results['collection_unmatched_cl'], 'Debit') + safe_sum(results['collection_unmatched_cl'], 'Credit')
    rev_cross_val    = safe_sum(results['reversal_cross_ledger'], 'VL Original Debit') + safe_sum(results['reversal_cross_ledger'], 'VL Original Credit')
    rev_int_val      = safe_sum(results['reversal_vl_internal'], 'VL Original Debit') + safe_sum(results['reversal_vl_internal'], 'VL Original Credit')
    rev_un_val       = safe_sum(results['reversal_unmatched'], 'VL Debit') + safe_sum(results['reversal_unmatched'], 'VL Credit')

    total_matched_cnt = inv_matched_cnt + dn_matched_cnt + cn_matched_cnt + col_matched_cnt
    total_matched_val = inv_matched_val + dn_matched_val + cn_matched_val + col_matched_val
    total_un_vl_cnt   = inv_un_vl_cnt + dn_un_vl_cnt + col_un_vl_cnt
    total_un_vl_val   = inv_un_vl_val + dn_un_vl_val + col_un_vl_val
    total_un_cl_cnt   = inv_un_cl_cnt + dn_un_cl_cnt + col_un_cl_cnt
    total_un_cl_val   = inv_un_cl_val + dn_un_cl_val + col_un_cl_val

    # ── DOWNLOAD CSS (styling only — button rendered after stats) ──
    st.markdown("""
    <style>
    div[data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #1a6b45, #00d4aa) !important;
        color: white !important;
        font-family: 'Syne', sans-serif !important;
        font-size: 1.1rem !important;
        font-weight: 800 !important;
        padding: 0.85rem 2rem !important;
        border-radius: 10px !important;
        border: none !important;
        letter-spacing: 0.04em !important;
        box-shadow: 0 4px 18px rgba(0,212,170,0.35) !important;
        transition: all 0.2s !important;
    }
    div[data-testid="stDownloadButton"] > button:hover {
        opacity: 0.9 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 24px rgba(0,212,170,0.5) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Build Excel once and cache in session_state to avoid re-running on every interaction
    if 'excel_data' not in st.session_state or st.session_state.get('excel_key') != id(results):
        try:
            st.session_state['excel_data'] = build_excel(results, vl_ann_df, cl_ann_df, VL, CL)
            st.session_state['excel_key'] = id(results)
        except Exception as e:
            st.session_state['excel_data'] = None
            st.error(f"Error generating Excel report: {e}")

    excel_data = st.session_state.get('excel_data')

    # ── TOP DOWNLOAD BUTTON ──
    if excel_data:
        st.download_button(
            label="⬇️  Download Ledger Reconciliation Report (.xlsx)",
            data=excel_data,
            file_name=f"Recon_{VL}_{CL}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="download_top",
        )

    st.markdown("---")
    st.markdown(f"""
    <div class="stat-grid">
        <div class="stat-card matched">
            <div class="stat-label">Total Matched</div>
            <div class="stat-value">{total_matched_cnt}</div>
            <div class="stat-sub">{fmt_inr(total_matched_val)}</div>
        </div>
        <div class="stat-card unmatched">
            <div class="stat-label">Unmatched ({VL})</div>
            <div class="stat-value">{total_un_vl_cnt}</div>
            <div class="stat-sub">{fmt_inr(total_un_vl_val)} · {CL}: {total_un_cl_cnt} ({fmt_inr(total_un_cl_val)})</div>
        </div>
        <div class="stat-card partial">
            <div class="stat-label">Credit Notes Matched</div>
            <div class="stat-value">{cn_matched_cnt}</div>
            <div class="stat-sub">{fmt_inr(cn_matched_val)} vs Discount DN / PRN</div>
        </div>
        <div class="stat-card total">
            <div class="stat-label">Invoice Matches</div>
            <div class="stat-value">{inv_matched_cnt}</div>
            <div class="stat-sub">{fmt_inr(inv_matched_val)} · DN: {dn_matched_cnt} · Coll: {col_matched_cnt}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── DASHBOARD SUMMARY TABLE ──
    st.markdown("### 📊 Reconciliation Summary")
    st.caption("Each row shows Count and Value. Click 'View ↓' to jump to that annexure.")

    inv_vl  = inv_matched_cnt + inv_un_vl_cnt
    inv_cl  = inv_matched_cnt + inv_un_cl_cnt
    dn_vl   = dn_matched_cnt  + dn_un_vl_cnt
    dn_cl   = dn_matched_cnt  + dn_un_cl_cnt
    col_vl  = col_matched_cnt + col_un_vl_cnt
    col_cl  = col_matched_cnt + col_un_cl_cnt
    cn_vl   = cn_matched_cnt  + len([r for r in results['invoice_unmatched_vl'] if is_credit_note(str(r.get('Type','')))])

    inv_pct = round(inv_matched_cnt / inv_vl * 100, 1) if inv_vl else 0
    dn_pct  = round(dn_matched_cnt  / dn_vl  * 100, 1) if dn_vl  else 0
    col_pct = round(col_matched_cnt / col_vl * 100, 1) if col_vl else 0
    cn_pct  = round(cn_matched_cnt  / cn_vl  * 100, 1) if cn_vl  else 0

    def sum_row(label, vl_cnt, matched_cnt, un_vl_cnt, vl_val, matched_val, un_vl_val,
                cl_cnt, un_cl_cnt, cl_val, un_cl_val, pct, tab_key, color,
                mis_cnt=None, miss_cnt=None, mis_val=None, miss_val=None):
        """Render one summary row with count+value and optional mismatch/missing split."""
        ca, cb, cc, cd, ce, cf, cg, ch = st.columns([2.8, 1.4, 1.4, 1.4, 1.4, 1.4, 1, 1.2])
        ca.markdown(f"**{label}**")
        cb.markdown(f"<div style='text-align:center;font-size:0.78rem'><b>{vl_cnt}</b><br><span style='color:var(--muted);font-size:0.68rem'>{fmt_inr(vl_val)}</span></div>", unsafe_allow_html=True)
        cc.markdown(f"<div style='text-align:center;font-size:0.78rem;color:{color}'><b>{matched_cnt}</b><br><span style='font-size:0.68rem'>{fmt_inr(matched_val)}</span></div>", unsafe_allow_html=True)
        # Unmatched VL split into mismatch + missing if provided
        if mis_cnt is not None:
            cd.markdown(f"<div style='text-align:center;font-size:0.72rem;color:#ff4d6d'><b>{un_vl_cnt}</b><br>Mis: {mis_cnt} | Miss: {miss_cnt}</div>", unsafe_allow_html=True)
        else:
            cd.markdown(f"<div style='text-align:center;font-size:0.78rem;color:#ff4d6d'><b>{un_vl_cnt}</b><br><span style='font-size:0.68rem'>{fmt_inr(un_vl_val)}</span></div>", unsafe_allow_html=True)
        ce.markdown(f"<div style='text-align:center;font-size:0.78rem'><b>{cl_cnt}</b></div>", unsafe_allow_html=True)
        cf.markdown(f"<div style='text-align:center;font-size:0.78rem;color:#ff4d6d'><b>{un_cl_cnt}</b><br><span style='font-size:0.68rem'>{fmt_inr(un_cl_val)}</span></div>", unsafe_allow_html=True)
        cg.markdown(f"<div style='text-align:center;color:{color};font-size:0.82rem'><b>{pct}%</b></div>", unsafe_allow_html=True)
        if ch.button("View ↓", key=f"btn_{tab_key}"):
            st.session_state['active_tab'] = tab_key
        st.markdown("<hr style='margin:3px 0;border-color:#252c3d'>", unsafe_allow_html=True)

    # Header row
    hc = st.columns([2.8, 1.4, 1.4, 1.4, 1.4, 1.4, 1, 1.2])
    for col_obj, label in zip(hc, [
        'Category',
        f'{VL} Total (Cnt/Val)',
        'Matched (Cnt/Val)',
        f'{VL} Unmatch\n(Mis|Miss)',
        f'{CL} Total',
        f'{CL} Unmatch (Val)',
        'Match%',
        'Annexure'
    ]):
        col_obj.markdown(f"<div style='font-size:0.65rem;color:#4f8eff;text-transform:uppercase;font-weight:700;text-align:center'>{label}</div>", unsafe_allow_html=True)
    st.markdown("<hr style='margin:4px 0;border-color:#4f8eff;border-width:2px'>", unsafe_allow_html=True)

    sum_row('🧾 Invoices', inv_vl, inv_matched_cnt, inv_un_vl_cnt,
            inv_matched_val+inv_un_vl_val, inv_matched_val, inv_un_vl_val,
            inv_cl, inv_un_cl_cnt, inv_matched_val+inv_un_cl_val, inv_un_cl_val,
            inv_pct, 'inv', '#00d4aa')

    sum_row('🟡 Credit Notes ↔ Disc DN/PRN', cn_vl, cn_matched_cnt, cn_vl-cn_matched_cnt,
            cn_matched_val, cn_matched_val, 0,
            '-', '-', 0, 0,
            cn_pct, 'cn_match', '#4f8eff')

    sum_row('📝 Debit Notes', dn_vl, dn_matched_cnt, dn_un_vl_cnt,
            dn_matched_val+dn_un_vl_val, dn_matched_val, dn_un_vl_val,
            dn_cl, dn_un_cl_cnt, dn_matched_val+dn_un_cl_val, dn_un_cl_val,
            dn_pct, 'dn', '#00d4aa')

    sum_row('💰 Collections', col_vl, col_matched_cnt, col_un_vl_cnt,
            col_matched_val+col_un_vl_val, col_matched_val, col_un_vl_val,
            col_cl, col_un_cl_cnt, col_matched_val+col_un_cl_val, col_un_cl_val,
            col_pct, 'col', '#00d4aa')

    sum_row('🔁 Reversal — Also in CL', rev_cross_cnt, rev_cross_cnt, 0,
            rev_cross_val, rev_cross_val, 0,
            rev_cross_cnt, 0, rev_cross_val, 0,
            100 if rev_cross_cnt else 0, 'rev_cross', '#ff8c42')

    sum_row('🔄 Reversal — Not in CL', rev_int_cnt*2, rev_int_cnt, 0,
            rev_int_val, rev_int_val, 0,
            '-', '-', 0, 0,
            100 if rev_int_cnt else 0, 'rev_int', '#00d4aa')

    sum_row('❓ Reversal — Mismatch/Missing', rev_un_cnt, 0, rev_un_cnt,
            rev_un_val, 0, rev_un_val,
            '-', '-', 0, 0,
            0, 'rev_un', '#ff4d6d',
            mis_cnt=rev_amt_mis_cnt, miss_cnt=rev_miss_cnt,
            mis_val=0, miss_val=0)

    # TOTAL row
    tc = st.columns([2.8, 1.4, 1.4, 1.4, 1.4, 1.4, 1, 1.2])
    tc[0].markdown("**TOTAL**")
    tc[1].markdown(f"<div style='text-align:center;font-weight:700'>{inv_vl+dn_vl+col_vl}<br><span style='font-size:0.7rem'>{fmt_inr(inv_matched_val+inv_un_vl_val+dn_matched_val+dn_un_vl_val+col_matched_val+col_un_vl_val)}</span></div>", unsafe_allow_html=True)
    tc[2].markdown(f"<div style='text-align:center;font-weight:700;color:#00d4aa'>{total_matched_cnt}<br><span style='font-size:0.7rem'>{fmt_inr(total_matched_val)}</span></div>", unsafe_allow_html=True)
    tc[3].markdown(f"<div style='text-align:center;font-weight:700;color:#ff4d6d'>{total_un_vl_cnt}<br><span style='font-size:0.7rem'>{fmt_inr(total_un_vl_val)}</span></div>", unsafe_allow_html=True)
    tc[4].markdown(f"<div style='text-align:center;font-weight:700'>{inv_cl+dn_cl+col_cl}</div>", unsafe_allow_html=True)
    tc[5].markdown(f"<div style='text-align:center;font-weight:700;color:#ff4d6d'>{total_un_cl_cnt}<br><span style='font-size:0.7rem'>{fmt_inr(total_un_cl_val)}</span></div>", unsafe_allow_html=True)

    # ── LEDGER RECONCILIATION STATEMENT (sample format from image) ──
    st.markdown("---")
    st.markdown("### 📋 Ledger Reconciliation Statement")
    vl_bal     = float(vl_closing_val) if vl_closing_val is not None else 0.0
    cl_bal_act = float(cl_closing_val) if cl_closing_val is not None else 0.0
    adj_inv_vl = inv_un_vl_val
    # Credit notes in VL not matched to CL (these reduce VL balance from CL perspective)
    adj_cn_vl  = safe_sum(results.get('cn_unmatched_vl', []), 'Debit') + safe_sum(results.get('cn_unmatched_vl', []), 'Credit')
    adj_dn_cl  = dn_un_cl_val
    adj_inv_cl = inv_un_cl_val
    adj_pay_vl = col_un_vl_val
    adj_pay_cl = col_un_cl_val
    net_bal_b  = vl_bal - adj_inv_vl + adj_cn_vl - adj_dn_cl + adj_inv_cl + adj_pay_vl - adj_pay_cl
    cl_balance = cl_bal_act
    diff_bc    = net_bal_b - cl_balance

    recon_rows = [
        ("Particular", "Amount", "col_header"),
        (f"Balance as per {VL} Books as on (A)", fmt_inr(vl_bal), "header"),
        (f"Less:  Tax invoice delivered but not available in {CL}", fmt_inr(-adj_inv_vl), "less"),
        (f"Add:   Credit note available in {VL} but not in {CL} books", fmt_inr(adj_cn_vl), "add"),
        (f"Less:  Debit notes available in {CL} but not in {VL}", fmt_inr(-adj_dn_cl), "less"),
        (f"Add:   Tax invoice delivered but not available in {VL}", fmt_inr(adj_inv_cl), "add"),
        (f"Add:   Payment not available in {CL}", fmt_inr(adj_pay_vl), "add"),
        (f"Less:  Payment not available in {VL}", fmt_inr(-adj_pay_cl), "less"),
        ("", "", "blank"),
        (f"Net Balance as per {VL} Books — B", fmt_inr(net_bal_b), "total"),
        ("", "", "blank"),
        (f"Balance as per {CL} Books — C", fmt_inr(cl_balance), "cl_total"),
        ("", "", "blank"),
        ("Unreconciled Difference B - C", fmt_inr(diff_bc), "diff"),
        ("(This value should be zero after all the adjustments)", "", "note"),
    ]

    rs1, rs2 = st.columns([3, 1])
    for label, amount, row_type in recon_rows:
        styles = {
            "col_header": ("#1c2130", "#4f8eff", True),
            "header":     ("#1a3a6b", "#ffffff", True),
            "less":       ("#1a1a2a", "#ff9999", False),
            "add":        ("#1a2a1a", "#99ffcc", False),
            "blank":      ("#0d0f14", "#0d0f14", False),
            "total":      ("#1a3a6b", "#ffffff", True),
            "cl_total":   ("#1a5a3a", "#ffffff", True),
            "diff":       ("#cc0000", "#ffffff", True),
            "note":       ("#cc0000", "#ffcccc", False),
        }
        bg, fg, bold = styles.get(row_type, ("#141720", "#e8ecf4", False))
        fw = "700" if bold else "400"
        with rs1:
            st.markdown(f"<div style='background:{bg};color:{fg};padding:7px 14px;border-bottom:1px solid #252c3d;font-size:0.83rem;font-weight:{fw}'>{label}</div>", unsafe_allow_html=True)
        with rs2:
            st.markdown(f"<div style='background:{bg};color:{fg};padding:7px 14px;border-bottom:1px solid #252c3d;font-size:0.83rem;font-weight:{fw};text-align:right'>{amount}</div>", unsafe_allow_html=True)

    # ── DOWNLOAD (bottom — secondary) ──
    st.markdown("---")
    if excel_data:
        st.download_button(
            label="⬇️  Download Reconciliation Report",
            data=excel_data,
            file_name=f"Recon_{VL}_{CL}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
            key="download_bottom",
        )
    st.markdown("---")

    # Determine active tab from summary button clicks
    tab_map = {'inv': 0, 'cn_match': 1, 'dn': 1, 'col': 2, 'rev_cross': 3, 'rev_int': 3, 'rev_un': 3}

    tabs = st.tabs([
        f"🧾 Invoices",
        f"📝 DN / Credit Notes",
        f"💰 Collections",
        f"🔁 Reversals",
        f"⚠️ All Unmatched",
        f"📘 {VL} Ledger",
        f"📗 {CL} Ledger",
    ])

    with tabs[0]:
        st.markdown(f'<a name="inv"></a>', unsafe_allow_html=True)
        st.markdown(f'<span class="section-tag tag-matched">MATCHED INVOICES — {VL} vs {CL}</span>', unsafe_allow_html=True)
        display_df(results['invoice_matched'])
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f'<span class="section-tag tag-vl">UNMATCHED — {VL} (VENDOR LEDGER)</span>', unsafe_allow_html=True)
            display_df(results['invoice_unmatched_vl'])
        with c2:
            st.markdown(f'<span class="section-tag tag-cl">UNMATCHED — {CL} (CUSTOMER LEDGER)</span>', unsafe_allow_html=True)
            display_df(results['invoice_unmatched_cl'])

    with tabs[1]:
        st.markdown('<a name="cn_match"></a>', unsafe_allow_html=True)
        st.markdown(f'<span class="section-tag tag-blue">🟡 CREDIT NOTES ({VL}) ↔ DISCOUNT DN / PRN ({CL})</span>', unsafe_allow_html=True)
        display_df(cn_matched)
        st.markdown("---")
        st.markdown(f'<span class="section-tag tag-matched">MATCHED DEBIT NOTES</span>', unsafe_allow_html=True)
        display_df(dn_only_matched)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f'<span class="section-tag tag-vl">UNMATCHED DEBIT NOTES — {VL}</span>', unsafe_allow_html=True)
            display_df(results['dn_unmatched_vl'])
        with c2:
            st.markdown(f'<span class="section-tag tag-cl">UNMATCHED DEBIT NOTES — {CL}</span>', unsafe_allow_html=True)
            display_df(results['dn_unmatched_cl'])

    with tabs[2]:
        st.markdown('<a name="col"></a>', unsafe_allow_html=True)
        st.markdown(f'<span class="section-tag tag-matched">MATCHED COLLECTIONS</span>', unsafe_allow_html=True)
        display_df(results['collection_matched'])
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f'<span class="section-tag tag-vl">UNMATCHED COLLECTIONS — {VL}</span>', unsafe_allow_html=True)
            display_df(results['collection_unmatched_vl'])
        with c2:
            st.markdown(f'<span class="section-tag tag-cl">UNMATCHED COLLECTIONS — {CL}</span>', unsafe_allow_html=True)
            display_df(results['collection_unmatched_cl'])

    with tabs[3]:
        st.markdown('<a name="rev_cross"></a>', unsafe_allow_html=True)
        st.markdown(f'<span class="section-tag tag-partial">⚠️ ANNEXURE A — REVERSED IN {VL} | INVOICE ALSO IN {CL}</span>', unsafe_allow_html=True)
        st.caption(f"These invoices were reversed in {VL} but the original invoice also exists in {CL}. Needs review.")
        display_df(results['reversal_cross_ledger'])
        st.markdown("---")
        st.markdown('<a name="rev_int"></a>', unsafe_allow_html=True)
        st.markdown(f'<span class="section-tag tag-matched">✅ ANNEXURE B — REVERSED IN {VL} | NOT IN {CL}</span>', unsafe_allow_html=True)
        display_df(results['reversal_vl_internal'])
        st.markdown("---")
        st.markdown('<a name="rev_un"></a>', unsafe_allow_html=True)
        # Split Annexure C into Amount Mismatch and Missing
        rev_mis = [r for r in results['reversal_unmatched'] if r.get('Reason','') == 'Amount Mismatch']
        rev_miss = [r for r in results['reversal_unmatched'] if r.get('Reason','') != 'Amount Mismatch']
        st.markdown(f'<span class="section-tag tag-unmatched">❌ ANNEXURE C1 — REVERSAL | AMOUNT MISMATCH ({len(rev_mis)} items)</span>', unsafe_allow_html=True)
        st.caption("Reversal found in VL but amounts do not match. VL Reversal Amount and Original Amount shown separately.")
        display_df(rev_mis)
        st.markdown(f'<span class="section-tag tag-unmatched">❌ ANNEXURE C2 — REVERSAL | ORIGINAL NOT FOUND ({len(rev_miss)} items)</span>', unsafe_allow_html=True)
        display_df(rev_miss)

    with tabs[4]:
        all_unmatched = []
        for item in results['invoice_unmatched_vl'] + results['dn_unmatched_vl'] + results['collection_unmatched_vl']:
            item = dict(item); item['Ledger'] = VL; all_unmatched.append(item)
        for item in results['invoice_unmatched_cl'] + results['dn_unmatched_cl'] + results['collection_unmatched_cl']:
            item = dict(item); item['Ledger'] = CL; all_unmatched.append(item)
        st.markdown(f'<span class="section-tag tag-unmatched">ALL UNMATCHED ITEMS — {VL} & {CL}</span>', unsafe_allow_html=True)
        display_df(all_unmatched)

    with tabs[5]:
        st.markdown(f'<span class="section-tag tag-vl">📘 {VL} — VENDOR LEDGER WITH REMARKS</span>', unsafe_allow_html=True)
        st.caption(f"🟢 Green = Matched | 🔴 Red = Unmatched | 🟡 Yellow = Reversal")
        if not vl_ann_df.empty:
            disp = vl_ann_df.copy()
            for col in disp.columns:
                if 'date' in col.lower():
                    disp[col] = pd.to_datetime(disp[col], errors='coerce').dt.strftime('%d-%b-%Y').fillna('')
            st.dataframe(disp, use_container_width=True, hide_index=True)
            if vl_closing_val:
                st.info(f"**{VL} Closing Balance: {fmt_inr(vl_closing_val)}**")

    with tabs[6]:
        st.markdown(f'<span class="section-tag tag-cl">📗 {CL} — CUSTOMER LEDGER WITH REMARKS</span>', unsafe_allow_html=True)
        st.caption(f"🟢 Green = Matched | 🔴 Red = Unmatched")
        if not cl_ann_df.empty:
            disp = cl_ann_df.copy()
            for col in disp.columns:
                if 'date' in col.lower():
                    disp[col] = pd.to_datetime(disp[col], errors='coerce').dt.strftime('%d-%b-%Y').fillna('')
            st.dataframe(disp, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
